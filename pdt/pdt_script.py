import os
import re
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── CONFIG ──────────────────────────────────────────────
EXCEL_FILE = "physical_design_results.xlsx"

# ── STEP 1: Pick PD report files ────────────────────────
def pick_files():
    print("\n=== VLSI Physical Design Report Parser ===")
    print("Parses: Floorplan, Placement, Routing, DRC reports\n")

    folder = input("Enter folder path with PD report files (or press Enter for current): ").strip()
    if not folder:
        folder = "."

    all_files = (
        glob.glob(os.path.join(folder, "*.rpt")) +
        glob.glob(os.path.join(folder, "*.log")) +
        glob.glob(os.path.join(folder, "*.txt"))
    )

    if not all_files:
        print(f"No report files found in: {folder}")
        return []

    print(f"\nFound {len(all_files)} file(s):\n")
    for i, f in enumerate(all_files):
        print(f"  [{i+1}] {os.path.basename(f)}")

    print("\n  A = All   or   1,2,3 = specific files")
    choice = input("\nYour choice: ").strip().upper()

    if choice == "A":
        return all_files
    selected = []
    for n in choice.split(","):
        n = n.strip()
        if n.isdigit() and 1 <= int(n) <= len(all_files):
            selected.append(all_files[int(n)-1])
    return selected

# ── STEP 2: Parse PD report ─────────────────────────────
def parse_pd(filepath):
    with open(filepath) as f:
        content = f.read()

    data = {
        # Floorplan
        "die_area"       : None,
        "core_area"       : None,
        "utilization"     : None,
        "aspect_ratio"    : None,
        # Placement
        "total_cells"     : None,
        "placed_cells"    : None,
        "unplaced_cells"  : None,
        "placement_density": None,
        # Routing
        "total_nets"      : None,
        "routed_nets"     : None,
        "unrouted_nets"   : None,
        "total_wire_length": None,
        "via_count"       : None,
        # DRC
        "drc_violations"  : None,
        "short_violations": None,
        "spacing_violations": None,
        "antenna_violations": None,
        # Congestion
        "overflow"        : None,
        "h_overflow"      : None,
        "v_overflow"      : None,
    }

    for line in content.splitlines():
        s = line.strip()

        # ── Floorplan ──
        m = re.search(r"Die\s+Area\s*[:\-]?\s*([\d.]+)", s, re.IGNORECASE)
        if m: data["die_area"] = float(m.group(1))

        m = re.search(r"Core\s+Area\s*[:\-]?\s*([\d.]+)", s, re.IGNORECASE)
        if m: data["core_area"] = float(m.group(1))

        m = re.search(r"[Uu]tilization\s*[:\-]?\s*([\d.]+)\s*%?", s)
        if m: data["utilization"] = float(m.group(1))

        m = re.search(r"[Aa]spect\s+[Rr]atio\s*[:\-]?\s*([\d.]+)", s)
        if m: data["aspect_ratio"] = float(m.group(1))

        # ── Placement ──
        m = re.search(r"Total\s+[Cc]ells?\s*[:\-]?\s*(\d+)", s)
        if m: data["total_cells"] = int(m.group(1))

        m = re.search(r"[Pp]laced\s+[Cc]ells?\s*[:\-]?\s*(\d+)", s)
        if m: data["placed_cells"] = int(m.group(1))

        m = re.search(r"[Uu]nplaced\s+[Cc]ells?\s*[:\-]?\s*(\d+)", s)
        if m: data["unplaced_cells"] = int(m.group(1))

        m = re.search(r"[Pp]lacement\s+[Dd]ensity\s*[:\-]?\s*([\d.]+)", s)
        if m: data["placement_density"] = float(m.group(1))

        # ── Routing ──
        m = re.search(r"Total\s+[Nn]ets?\s*[:\-]?\s*(\d+)", s)
        if m: data["total_nets"] = int(m.group(1))

        m = re.search(r"[Rr]outed\s+[Nn]ets?\s*[:\-]?\s*(\d+)", s)
        if m: data["routed_nets"] = int(m.group(1))

        m = re.search(r"[Uu]n[Rr]outed\s+[Nn]ets?\s*[:\-]?\s*(\d+)", s)
        if m: data["unrouted_nets"] = int(m.group(1))

        m = re.search(r"[Tt]otal\s+[Ww]ire\s+[Ll]ength\s*[:\-]?\s*([\d.]+)", s)
        if m: data["total_wire_length"] = float(m.group(1))

        m = re.search(r"[Vv]ia\s+[Cc]ount\s*[:\-]?\s*(\d+)", s)
        if m: data["via_count"] = int(m.group(1))

        # ── DRC ──
        m = re.search(r"(Total\s+)?DRC\s+[Vv]iolations?\s*[:\-]?\s*(\d+)", s)
        if m: data["drc_violations"] = int(m.group(2))

        m = re.search(r"[Ss]hort\s+[Vv]iolations?\s*[:\-]?\s*(\d+)", s)
        if m: data["short_violations"] = int(m.group(1))

        m = re.search(r"[Ss]pacing\s+[Vv]iolations?\s*[:\-]?\s*(\d+)", s)
        if m: data["spacing_violations"] = int(m.group(1))

        m = re.search(r"[Aa]ntenna\s+[Vv]iolations?\s*[:\-]?\s*(\d+)", s)
        if m: data["antenna_violations"] = int(m.group(1))

        # ── Congestion / Overflow ──
        m = re.search(r"(Total\s+)?[Oo]verflow\s*[:\-]?\s*(\d+)", s)
        if m: data["overflow"] = int(m.group(2))

        m = re.search(r"[Hh]orizontal\s+[Oo]verflow\s*[:\-]?\s*(\d+)", s)
        if m: data["h_overflow"] = int(m.group(1))

        m = re.search(r"[Vv]ertical\s+[Oo]verflow\s*[:\-]?\s*(\d+)", s)
        if m: data["v_overflow"] = int(m.group(1))

    # Compute unplaced / unrouted if missing
    if data["unplaced_cells"] is None and data["total_cells"] and data["placed_cells"]:
        data["unplaced_cells"] = data["total_cells"] - data["placed_cells"]
    if data["unrouted_nets"] is None and data["total_nets"] and data["routed_nets"]:
        data["unrouted_nets"] = data["total_nets"] - data["routed_nets"]

    return data

# ── STEP 3: Print terminal summary ──────────────────────
def print_terminal(fname, data):
    drc_ok = data["drc_violations"] == 0
    print(f"\n{'='*55}")
    print(f"  File          : {fname}")
    print(f"  Die Area      : {data['die_area']} um²")
    print(f"  Core Area     : {data['core_area']} um²")
    print(f"  Utilization   : {data['utilization']}%")
    print(f"  Total Cells   : {data['total_cells']}  Placed: {data['placed_cells']}  Unplaced: {data['unplaced_cells']}")
    print(f"  Total Nets    : {data['total_nets']}  Routed: {data['routed_nets']}  Unrouted: {data['unrouted_nets']}")
    print(f"  Wire Length   : {data['total_wire_length']} um")
    print(f"  Via Count     : {data['via_count']}")
    drc_str = "✅ CLEAN" if drc_ok else f"❌ {data['drc_violations']} violations"
    print(f"  DRC           : {drc_str}")
    print(f"  Overflow      : {data['overflow']}")
    print(f"{'='*55}")

# ── STEP 4: Save to Excel ────────────────────────────────
def save_excel(all_results):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR = PatternFill("solid", start_color="1F3864")
    ALT = PatternFill("solid", start_color="F2F2F2")
    WHT = PatternFill("solid", start_color="FFFFFF")
    GRN = PatternFill("solid", start_color="C6EFCE")
    RED = PatternFill("solid", start_color="FFC7CE")
    YLW = PatternFill("solid", start_color="FFEB9C")
    BLU = PatternFill("solid", start_color="DDEEFF")

    def hdr(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill      = HDR
        c.alignment = Alignment(horizontal="center")

    def rf(r): return ALT if r % 2 == 0 else WHT

    def sc(ws, r, c, v, fill=None, bold=False, center=False):
        cell = ws.cell(r, c, v)
        cell.font = Font(bold=bold, name="Arial")
        if fill: cell.fill = fill
        if center: cell.alignment = Alignment(horizontal="center")

    def util_fill(v):
        if v is None: return WHT
        if v > 90: return RED
        if v > 75: return YLW
        return GRN

    # ── Sheet 1: Floorplan ──
    ws1 = wb.create_sheet("Floorplan")
    for ci, h in enumerate(["File","Die Area (um²)","Core Area (um²)","Utilization %","Aspect Ratio"], 1):
        hdr(ws1, 1, ci, h)
    for col, w in zip("ABCDE", [28,16,16,14,13]):
        ws1.column_dimensions[col].width = w

    for ri, (fname, data) in enumerate(all_results, 2):
        f = rf(ri)
        sc(ws1, ri, 1, fname,               f)
        sc(ws1, ri, 2, data["die_area"],    BLU, center=True)
        sc(ws1, ri, 3, data["core_area"],   BLU, center=True)
        sc(ws1, ri, 4, data["utilization"], util_fill(data["utilization"]), bold=True, center=True)
        sc(ws1, ri, 5, data["aspect_ratio"],f, center=True)

    # ── Sheet 2: Placement ──
    ws2 = wb.create_sheet("Placement")
    for ci, h in enumerate(["File","Total Cells","Placed","Unplaced","Density %"], 1):
        hdr(ws2, 1, ci, h)
    for col, w in zip("ABCDE", [28,12,12,12,12]):
        ws2.column_dimensions[col].width = w

    for ri, (fname, data) in enumerate(all_results, 2):
        f = rf(ri)
        unp = data["unplaced_cells"] or 0
        sc(ws2, ri, 1, fname,                   f)
        sc(ws2, ri, 2, data["total_cells"],      f, center=True)
        sc(ws2, ri, 3, data["placed_cells"],     GRN, center=True)
        sc(ws2, ri, 4, unp,                      RED if unp else GRN, bold=bool(unp), center=True)
        sc(ws2, ri, 5, data["placement_density"],f, center=True)

    # ── Sheet 3: Routing ──
    ws3 = wb.create_sheet("Routing")
    for ci, h in enumerate(["File","Total Nets","Routed","Unrouted","Wire Length (um)","Via Count"], 1):
        hdr(ws3, 1, ci, h)
    for col, w in zip("ABCDEF", [28,12,12,12,17,12]):
        ws3.column_dimensions[col].width = w

    for ri, (fname, data) in enumerate(all_results, 2):
        f  = rf(ri)
        unr = data["unrouted_nets"] or 0
        sc(ws3, ri, 1, fname,                    f)
        sc(ws3, ri, 2, data["total_nets"],        f, center=True)
        sc(ws3, ri, 3, data["routed_nets"],       GRN, center=True)
        sc(ws3, ri, 4, unr,                       RED if unr else GRN, bold=bool(unr), center=True)
        sc(ws3, ri, 5, data["total_wire_length"], BLU, center=True)
        sc(ws3, ri, 6, data["via_count"],         f, center=True)

    # ── Sheet 4: DRC & Congestion ──
    ws4 = wb.create_sheet("DRC & Congestion")
    for ci, h in enumerate(["File","DRC Total","Shorts","Spacing","Antenna","Overflow","H-Overflow","V-Overflow","Status"], 1):
        hdr(ws4, 1, ci, h)
    for col, w in zip("ABCDEFGHI", [28,10,10,10,10,10,12,12,10]):
        ws4.column_dimensions[col].width = w

    for ri, (fname, data) in enumerate(all_results, 2):
        f = rf(ri)
        drc = data["drc_violations"] or 0
        ovf = data["overflow"] or 0
        status = "CLEAN" if drc == 0 and ovf == 0 else "VIOLATIONS"
        sf = GRN if status == "CLEAN" else RED
        sc(ws4, ri, 1, fname,                       f)
        sc(ws4, ri, 2, drc,                         RED if drc else GRN, bold=bool(drc), center=True)
        sc(ws4, ri, 3, data["short_violations"],     RED if data["short_violations"] else f, center=True)
        sc(ws4, ri, 4, data["spacing_violations"],   RED if data["spacing_violations"] else f, center=True)
        sc(ws4, ri, 5, data["antenna_violations"],   YLW if data["antenna_violations"] else f, center=True)
        sc(ws4, ri, 6, ovf,                          RED if ovf else GRN, bold=bool(ovf), center=True)
        sc(ws4, ri, 7, data["h_overflow"],           f, center=True)
        sc(ws4, ri, 8, data["v_overflow"],           f, center=True)
        sc(ws4, ri, 9, status,                       sf, bold=True, center=True)

    wb.save(EXCEL_FILE)
    print(f"\nExcel saved : {EXCEL_FILE}")
    print(f"  Sheet 1 → Floorplan   (die, core area, utilization)")
    print(f"  Sheet 2 → Placement   (cells placed, unplaced, density)")
    print(f"  Sheet 3 → Routing     (nets, wire length, vias)")
    print(f"  Sheet 4 → DRC         (violations, congestion, overflow)")

# ── MAIN ────────────────────────────────────────────────
def main():
    files = pick_files()
    if not files:
        print("No files selected. Exiting.")
        return

    all_results = []
    for filepath in files:
        fname = os.path.basename(filepath)
        print(f"\nParsing: {fname} ...")
        data = parse_pd(filepath)
        print_terminal(fname, data)
        all_results.append((fname, data))

    save_excel(all_results)

if __name__ == "__main__":
    main()
