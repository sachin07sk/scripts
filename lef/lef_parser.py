import os
import re
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── CONFIG ──────────────────────────────────────────────
EXCEL_FILE = "lef_def_results.xlsx"

# ── STEP 1: Ask user to pick files ──────────────────────
def pick_files():
    print("\n=== VLSI LEF / DEF Parser ===\n")

    folder = input("Enter folder path with LEF/DEF files (or press Enter for current): ").strip()
    if not folder:
        folder = "."

    lef_files = glob.glob(os.path.join(folder, "*.lef"))
    def_files = glob.glob(os.path.join(folder, "*.def"))

    print(f"\nFound: {len(lef_files)} LEF file(s)  |  {len(def_files)} DEF file(s)\n")

    def select(files, ftype):
        if not files:
            print(f"  No {ftype} files found.")
            return []
        for i, f in enumerate(files):
            print(f"  [{i+1}] {os.path.basename(f)}")
        print(f"  A = All {ftype}   or   1,2,3 = specific")
        choice = input(f"  Pick {ftype} files: ").strip().upper()
        if choice == "A":
            return files
        selected = []
        for n in choice.split(","):
            n = n.strip()
            if n.isdigit() and 1 <= int(n) <= len(files):
                selected.append(files[int(n)-1])
        return selected

    print("-- LEF files --")
    chosen_lef = select(lef_files, "LEF")
    print("\n-- DEF files --")
    chosen_def = select(def_files, "DEF")

    return chosen_lef, chosen_def

# ── STEP 2: Parse LEF file ───────────────────────────────
def parse_lef(filepath):
    with open(filepath) as f:
        lines = f.readlines()

    macros = []   # cells
    layers = []
    current_macro = None
    current_pin   = None
    in_pin        = False

    for line in lines:
        line = line.strip()

        # ── Layer ──
        m = re.match(r"^LAYER\s+(\S+)", line)
        if m and "END" not in line:
            layers.append({"layer_name": m.group(1), "type": "", "pitch": "", "width": ""})

        if layers:
            m = re.match(r"TYPE\s+(\S+)", line)
            if m:
                layers[-1]["type"] = m.group(1).rstrip(";")
            m = re.match(r"PITCH\s+([\d.]+)", line)
            if m:
                layers[-1]["pitch"] = m.group(1)
            m = re.match(r"WIDTH\s+([\d.]+)", line)
            if m:
                layers[-1]["width"] = m.group(1)

        # ── Macro (cell) start ──
        m = re.match(r"^MACRO\s+(\S+)", line)
        if m:
            current_macro = {
                "macro_name": m.group(1),
                "class"     : "",
                "width"     : "",
                "height"    : "",
                "pins"      : []
            }
            macros.append(current_macro)
            in_pin = False

        if current_macro is None:
            continue

        # ── Macro class ──
        m = re.match(r"CLASS\s+(\S+)", line)
        if m:
            current_macro["class"] = m.group(1).rstrip(";")

        # ── Macro size ──
        m = re.match(r"SIZE\s+([\d.]+)\s+BY\s+([\d.]+)", line)
        if m:
            current_macro["width"]  = m.group(1)
            current_macro["height"] = m.group(2)

        # ── Pin start ──
        m = re.match(r"^PIN\s+(\S+)", line)
        if m:
            current_pin = {
                "pin_name"  : m.group(1),
                "direction" : "",
                "use"       : "",
                "layer"     : "",
                "rect"      : ""
            }
            current_macro["pins"].append(current_pin)
            in_pin = True

        if in_pin and current_pin:
            m = re.match(r"DIRECTION\s+(\S+)", line)
            if m:
                current_pin["direction"] = m.group(1).rstrip(";")
            m = re.match(r"USE\s+(\S+)", line)
            if m:
                current_pin["use"] = m.group(1).rstrip(";")
            m = re.match(r"LAYER\s+(\S+)", line)
            if m:
                current_pin["layer"] = m.group(1).rstrip(";")
            m = re.search(r"RECT\s+([\d.\-]+)\s+([\d.\-]+)\s+([\d.\-]+)\s+([\d.\-]+)", line)
            if m:
                current_pin["rect"] = f"{m.group(1)} {m.group(2)} {m.group(3)} {m.group(4)}"

        if line.startswith("END PIN"):
            in_pin = False

        if re.match(r"^END\s+MACRO", line) or (line.startswith("END ") and current_macro and line == f"END {current_macro['macro_name']}"):
            current_macro = None

    return macros, layers

# ── STEP 3: Parse DEF file ───────────────────────────────
def parse_def(filepath):
    with open(filepath) as f:
        lines = f.readlines()

    components = []   # placed cells
    nets        = []
    pins        = []
    design_name = ""
    die_area    = ""
    in_comps    = False
    in_nets     = False
    in_pins     = False
    current_net = None

    for line in lines:
        line = line.strip()

        m = re.match(r"DESIGN\s+(\S+)", line)
        if m:
            design_name = m.group(1).rstrip(";")

        m = re.search(r"DIEAREA\s+\(\s*([\d.\-]+)\s+([\d.\-]+)\s*\)\s+\(\s*([\d.\-]+)\s+([\d.\-]+)\s*\)", line)
        if m:
            die_area = f"({m.group(1)},{m.group(2)}) ({m.group(3)},{m.group(4)})"

        # ── Components (placed cells) ──
        if line.startswith("COMPONENTS"):
            in_comps = True
        if line.startswith("END COMPONENTS"):
            in_comps = False
        if in_comps:
            # Example: - U_ADD MACRO_AND2 + PLACED ( 100 200 ) N ;
            m = re.match(r"^-\s+(\S+)\s+(\S+)", line)
            if m:
                comp = {"inst_name": m.group(1), "cell_type": m.group(2),
                        "x": "", "y": "", "orient": ""}
                pm = re.search(r"PLACED\s+\(\s*([\d.\-]+)\s+([\d.\-]+)\s*\)\s+(\S+)", line)
                if pm:
                    comp["x"] = pm.group(1)
                    comp["y"] = pm.group(2)
                    comp["orient"] = pm.group(3).rstrip(";")
                components.append(comp)

        # ── Pins ──
        if line.startswith("PINS"):
            in_pins = True
        if line.startswith("END PINS"):
            in_pins = False
        if in_pins:
            m = re.match(r"^-\s+(\S+)\s+\+\s+NET\s+(\S+)", line)
            if m:
                pin = {"pin_name": m.group(1), "net": m.group(2),
                       "direction": "", "layer": "", "x": "", "y": ""}
                dm = re.search(r"DIRECTION\s+(\S+)", line)
                if dm:
                    pin["direction"] = dm.group(1).rstrip(";")
                lm = re.search(r"LAYER\s+(\S+)\s+\(\s*([\d.\-]+)\s+([\d.\-]+)", line)
                if lm:
                    pin["layer"] = lm.group(1)
                    pin["x"]     = lm.group(2)
                    pin["y"]     = lm.group(3)
                pins.append(pin)

        # ── Nets ──
        if line.startswith("NETS"):
            in_nets = True
        if line.startswith("END NETS"):
            in_nets = False
            current_net = None
        if in_nets:
            m = re.match(r"^-\s+(\S+)", line)
            if m and not line.startswith("- ("):
                current_net = {"net_name": m.group(1), "connections": []}
                nets.append(current_net)
            if current_net:
                for conn in re.findall(r"\(\s*(\S+)\s+(\S+)\s*\)", line):
                    current_net["connections"].append(f"{conn[0]}/{conn[1]}")

    return design_name, die_area, components, pins, nets

# ── STEP 4: Print terminal summary ──────────────────────
def print_terminal(lef_data, def_data):
    for fname, macros, layers in lef_data:
        total_pins = sum(len(m["pins"]) for m in macros)
        print(f"\n{'='*50}")
        print(f"  LEF : {fname}")
        print(f"  Layers : {len(layers)}   Macros : {len(macros)}   Pins : {total_pins}")
        for l in layers[:5]:
            print(f"    Layer {l['layer_name']:<15} type={l['type']:<10} width={l['width']}")
        for m in macros[:3]:
            print(f"    Macro {m['macro_name']:<20} {m['width']} x {m['height']}  pins={len(m['pins'])}")

    for fname, design, die, comps, pins, nets in def_data:
        print(f"\n{'='*50}")
        print(f"  DEF    : {fname}")
        print(f"  Design : {design}")
        print(f"  Die    : {die}")
        print(f"  Components : {len(comps)}   Pins : {len(pins)}   Nets : {len(nets)}")

# ── STEP 5: Save to Excel ────────────────────────────────
def save_excel(lef_data, def_data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    HDR  = PatternFill("solid", start_color="1F3864")
    ALT  = PatternFill("solid", start_color="F2F2F2")
    WHT  = PatternFill("solid", start_color="FFFFFF")
    GRN  = PatternFill("solid", start_color="C6EFCE")
    BLU  = PatternFill("solid", start_color="DDEEFF")

    def hdr(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill      = HDR
        c.alignment = Alignment(horizontal="center")

    def row_fill(r):
        return ALT if r % 2 == 0 else WHT

    # ── Sheet: LEF Layers ──
    ws = wb.create_sheet("LEF Layers")
    for ci, h in enumerate(["File","Layer Name","Type","Pitch","Width"], 1):
        hdr(ws, 1, ci, h)
    for col, w in zip("ABCDE", [25,20,15,12,12]):
        ws.column_dimensions[col].width = w
    r = 2
    for fname, macros, layers in lef_data:
        for l in layers:
            f = row_fill(r)
            for ci, v in enumerate([fname, l["layer_name"], l["type"], l["pitch"], l["width"]], 1):
                ws.cell(r, ci, v).fill = f
                ws.cell(r, ci).font = Font(name="Arial")
            r += 1

    # ── Sheet: LEF Macros (cells) ──
    ws = wb.create_sheet("LEF Macros")
    for ci, h in enumerate(["File","Macro Name","Class","Width","Height","Pin Count"], 1):
        hdr(ws, 1, ci, h)
    for col, w in zip("ABCDEF", [25,25,15,10,10,10]):
        ws.column_dimensions[col].width = w
    r = 2
    for fname, macros, layers in lef_data:
        for m in macros:
            f = row_fill(r)
            for ci, v in enumerate([fname, m["macro_name"], m["class"], m["width"], m["height"], len(m["pins"])], 1):
                ws.cell(r, ci, v).fill = f
                ws.cell(r, ci).font = Font(name="Arial")
            r += 1

    # ── Sheet: LEF Pins ──
    ws = wb.create_sheet("LEF Pins")
    for ci, h in enumerate(["File","Macro","Pin Name","Direction","Use","Layer","Rect"], 1):
        hdr(ws, 1, ci, h)
    for col, w in zip("ABCDEFG", [25,25,20,12,12,15,35]):
        ws.column_dimensions[col].width = w
    r = 2
    for fname, macros, layers in lef_data:
        for m in macros:
            for p in m["pins"]:
                f = row_fill(r)
                for ci, v in enumerate([fname, m["macro_name"], p["pin_name"], p["direction"], p["use"], p["layer"], p["rect"]], 1):
                    ws.cell(r, ci, v).fill = f
                    ws.cell(r, ci).font = Font(name="Arial")
                r += 1

    # ── Sheet: DEF Components ──
    ws = wb.create_sheet("DEF Components")
    for ci, h in enumerate(["File","Design","Instance Name","Cell Type","X","Y","Orient"], 1):
        hdr(ws, 1, ci, h)
    for col, w in zip("ABCDEFG", [25,20,30,20,12,12,10]):
        ws.column_dimensions[col].width = w
    r = 2
    for fname, design, die, comps, pins, nets in def_data:
        for c in comps:
            f = row_fill(r)
            for ci, v in enumerate([fname, design, c["inst_name"], c["cell_type"], c["x"], c["y"], c["orient"]], 1):
                ws.cell(r, ci, v).fill = f
                ws.cell(r, ci).font = Font(name="Arial")
            r += 1

    # ── Sheet: DEF Pins ──
    ws = wb.create_sheet("DEF Pins")
    for ci, h in enumerate(["File","Design","Pin Name","Net","Direction","Layer","X","Y"], 1):
        hdr(ws, 1, ci, h)
    for col, w in zip("ABCDEFGH", [25,20,20,20,12,12,12,12]):
        ws.column_dimensions[col].width = w
    r = 2
    for fname, design, die, comps, pins, nets in def_data:
        for p in pins:
            f = row_fill(r)
            for ci, v in enumerate([fname, design, p["pin_name"], p["net"], p["direction"], p["layer"], p["x"], p["y"]], 1):
                ws.cell(r, ci, v).fill = f
                ws.cell(r, ci).font = Font(name="Arial")
            r += 1

    # ── Sheet: DEF Nets ──
    ws = wb.create_sheet("DEF Nets")
    for ci, h in enumerate(["File","Design","Net Name","Connection Count","Connections"], 1):
        hdr(ws, 1, ci, h)
    for col, w in zip("ABCDE", [25,20,25,18,60]):
        ws.column_dimensions[col].width = w
    r = 2
    for fname, design, die, comps, pins, nets in def_data:
        for n in nets:
            f = row_fill(r)
            conns = "  |  ".join(n["connections"][:10])
            if len(n["connections"]) > 10:
                conns += f"  ... +{len(n['connections'])-10} more"
            for ci, v in enumerate([fname, design, n["net_name"], len(n["connections"]), conns], 1):
                ws.cell(r, ci, v).fill = f
                ws.cell(r, ci).font = Font(name="Arial")
            r += 1

    wb.save(EXCEL_FILE)
    print(f"\nExcel saved : {EXCEL_FILE}")
    print(f"  Sheet 1 → LEF Layers")
    print(f"  Sheet 2 → LEF Macros  (cells + sizes)")
    print(f"  Sheet 3 → LEF Pins    (pin direction, layer, rect)")
    print(f"  Sheet 4 → DEF Components (placed instances + XY)")
    print(f"  Sheet 5 → DEF Pins    (IO pins + net)")
    print(f"  Sheet 6 → DEF Nets    (net connections)")

# ── MAIN ────────────────────────────────────────────────
def main():
    chosen_lef, chosen_def = pick_files()

    if not chosen_lef and not chosen_def:
        print("No files selected. Exiting.")
        return

    lef_data = []
    for f in chosen_lef:
        fname = os.path.basename(f)
        print(f"\nParsing LEF: {fname} ...")
        macros, layers = parse_lef(f)
        print(f"  → {len(layers)} layers, {len(macros)} macros, {sum(len(m['pins']) for m in macros)} pins")
        lef_data.append((fname, macros, layers))

    def_data = []
    for f in chosen_def:
        fname = os.path.basename(f)
        print(f"\nParsing DEF: {fname} ...")
        design, die, comps, pins, nets = parse_def(f)
        print(f"  → design={design}  components={len(comps)}  pins={len(pins)}  nets={len(nets)}")
        def_data.append((fname, design, die, comps, pins, nets))

    print_terminal(lef_data, def_data)
    save_excel(lef_data, def_data)

if __name__ == "__main__":
    main()
