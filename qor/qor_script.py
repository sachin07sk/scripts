import os
import re
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference

# ── CONFIG ──────────────────────────────────────────────
EXCEL_FILE = "qor_dashboard.xlsx"

# ── STEP 1: Pick QoR report files ───────────────────────
def pick_files():
    print("\n=== VLSI QoR Tracking Dashboard ===")
    print("    Tracks: Timing + Power + Area")
    print("    Compares: Across runs + corners\n")

    folder = input("Enter folder with QoR report files (or press Enter for current): ").strip()
    if not folder:
        folder = "."

    all_files = (
        glob.glob(os.path.join(folder, "*.rpt")) +
        glob.glob(os.path.join(folder, "*.log")) +
        glob.glob(os.path.join(folder, "*.txt"))
    )

    if not all_files:
        print(f"No report files found in: {folder}")
        return []

    print(f"\nFound {len(all_files)} file(s):\n")
    for i, f in enumerate(all_files):
        print(f"  [{i+1}] {os.path.basename(f)}")

    print("\n  A = All   or   1,2,3 = specific files")
    choice = input("\nYour choice: ").strip().upper()

    if choice == "A":
        return all_files
    selected = []
    for n in choice.split(","):
        n = n.strip()
        if n.isdigit() and 1 <= int(n) <= len(all_files):
            selected.append(all_files[int(n)-1])
    return selected

# ── STEP 2: Ask run info for each file ──────────────────
def get_run_info(fname):
    print(f"\n  File: {fname}")
    run    = input("    Run name   (e.g. run1, baseline): ").strip() or "run1"
    corner = input("    Corner     (e.g. ss_0v8_125c)  : ").strip() or "typical"
    return run, corner

# ── STEP 3: Parse QoR report ────────────────────────────
def parse_qor(filepath):
    with open(filepath) as f:
        content = f.read()

    data = {
        # Timing
        "wns"         : None,
        "tns"         : None,
        "violated"    : None,
        # Power
        "total_power" : None,
        "dynamic"     : None,
        "leakage"     : None,
        # Area
        "cell_count"  : None,
        "total_area"  : None,
        "utilization" : None,
    }

    # ── Timing ──
    m = re.search(r"WNS\s*[:\-]?\s*([-\d.]+)", content, re.IGNORECASE)
    if m: data["wns"] = float(m.group(1))

    m = re.search(r"TNS\s*[:\-]?\s*([-\d.]+)", content, re.IGNORECASE)
    if m: data["tns"] = float(m.group(1))

    m = re.search(r"(Violated|Failing)\s+[Pp]aths?\s*[:\-]?\s*(\d+)", content, re.IGNORECASE)
    if m: data["violated"] = int(m.group(2))

    # Also catch slack (VIOLATED) count
    if data["violated"] is None:
        data["violated"] = len(re.findall(r"slack\s*\(VIOLATED\)", content))

    # ── Power ──
    m = re.search(r"Total\s+Power\s*[:\-]?\s*([\d.eE+\-]+)", content, re.IGNORECASE)
    if m: data["total_power"] = float(m.group(1))

    m = re.search(r"(Switching|Dynamic)\s+Power\s*[:\-]?\s*([\d.eE+\-]+)", content, re.IGNORECASE)
    if m: data["dynamic"] = float(m.group(2))

    m = re.search(r"(Leakage|Static)\s+Power\s*[:\-]?\s*([\d.eE+\-]+)", content, re.IGNORECASE)
    if m: data["leakage"] = float(m.group(2))

    # ── Area ──
    m = re.search(r"(Total\s+)?Cell\s+Count\s*[:\-]?\s*(\d+)", content, re.IGNORECASE)
    if m: data["cell_count"] = int(m.group(2))

    m = re.search(r"Total\s+Area\s*[:\-]?\s*([\d.]+)", content, re.IGNORECASE)
    if m: data["total_area"] = float(m.group(1))

    m = re.search(r"Utilization\s*[:\-]?\s*([\d.]+)\s*%?", content, re.IGNORECASE)
    if m: data["utilization"] = float(m.group(1))

    return data

# ── STEP 4: Print terminal summary ──────────────────────
def print_terminal(records):
    print(f"\n{'='*70}")
    print(f"  {'Run':<12} {'Corner':<18} {'WNS':>8} {'TNS':>10} {'Power':>10} {'Area':>10} {'Util%':>7}")
    print(f"  {'-'*65}")
    for r in records:
        wns   = f"{r['wns']}"   if r['wns']   is not None else "-"
        tns   = f"{r['tns']}"   if r['tns']   is not None else "-"
        pwr   = f"{r['total_power']}" if r['total_power'] is not None else "-"
        area  = f"{r['total_area']}"  if r['total_area']  is not None else "-"
        util  = f"{r['utilization']}%" if r['utilization'] is not None else "-"
        print(f"  {r['run']:<12} {r['corner']:<18} {wns:>8} {tns:>10} {pwr:>10} {area:>10} {util:>7}")
    print(f"{'='*70}")

# ── STEP 5: Save Excel dashboard ────────────────────────
def save_excel(records):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR  = PatternFill("solid", start_color="1F3864")
    ALT  = PatternFill("solid", start_color="F2F2F2")
    WHT  = PatternFill("solid", start_color="FFFFFF")
    RED  = PatternFill("solid", start_color="FFC7CE")
    GRN  = PatternFill("solid", start_color="C6EFCE")
    YLW  = PatternFill("solid", start_color="FFEB9C")
    BLU  = PatternFill("solid", start_color="DDEEFF")

    def hdr(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill      = HDR
        c.alignment = Alignment(horizontal="center")

    def rf(r): return ALT if r % 2 == 0 else WHT

    def set_cell(ws, r, c, v, fill=None, bold=False, center=False):
        cell = ws.cell(r, c, v)
        cell.font = Font(bold=bold, name="Arial")
        if fill: cell.fill = fill
        if center: cell.alignment = Alignment(horizontal="center")
        return cell

    # ────────────────────────────────────────────────────
    # Sheet 1: Full QoR Table
    # ────────────────────────────────────────────────────
    ws1 = wb.create_sheet("QoR Table")
    cols = ["Run", "Corner", "File",
            "WNS", "TNS", "Violated Paths",
            "Total Power", "Dynamic", "Leakage",
            "Cell Count", "Total Area", "Utilization %"]
    for ci, h in enumerate(cols, 1):
        hdr(ws1, 1, ci, h)
    for col, w in zip("ABCDEFGHIJKL", [14,18,28,10,10,14,14,12,12,12,12,14]):
        ws1.column_dimensions[col].width = w

    for ri, r in enumerate(records, 2):
        bg = rf(ri)
        # timing fill
        wns_fill = RED if (r["wns"] is not None and r["wns"] < 0) else GRN if r["wns"] is not None else bg
        set_cell(ws1, ri, 1,  r["run"],          bg)
        set_cell(ws1, ri, 2,  r["corner"],        bg)
        set_cell(ws1, ri, 3,  r["file"],          bg)
        set_cell(ws1, ri, 4,  r["wns"],           wns_fill, bold=True, center=True)
        set_cell(ws1, ri, 5,  r["tns"],           wns_fill, center=True)
        set_cell(ws1, ri, 6,  r["violated"],      RED if r["violated"] else GRN, center=True)
        set_cell(ws1, ri, 7,  r["total_power"],   BLU, bold=True)
        set_cell(ws1, ri, 8,  r["dynamic"],       bg)
        set_cell(ws1, ri, 9,  r["leakage"],       bg)
        set_cell(ws1, ri, 10, r["cell_count"],    bg)
        set_cell(ws1, ri, 11, r["total_area"],    bg)
        util_fill = RED if (r["utilization"] and r["utilization"] > 90) else YLW if (r["utilization"] and r["utilization"] > 75) else GRN if r["utilization"] else bg
        set_cell(ws1, ri, 12, r["utilization"],   util_fill, center=True)

    # ────────────────────────────────────────────────────
    # Sheet 2: Compare Across Runs (group by corner)
    # ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Compare Runs")
    corners = sorted(set(r["corner"] for r in records))
    run_row = 1
    for corner in corners:
        corner_records = [r for r in records if r["corner"] == corner]
        # Section header
        ws2.cell(run_row, 1, f"Corner: {corner}").font = Font(bold=True, name="Arial", size=12, color="1F3864")
        run_row += 1
        for ci, h in enumerate(["Run", "WNS", "TNS", "Total Power", "Cell Count", "Utilization %"], 1):
            hdr(ws2, run_row, ci, h)
        run_row += 1
        for r in corner_records:
            bg = rf(run_row)
            wf = RED if (r["wns"] is not None and r["wns"] < 0) else GRN if r["wns"] is not None else bg
            set_cell(ws2, run_row, 1, r["run"],          bg)
            set_cell(ws2, run_row, 2, r["wns"],          wf, bold=True, center=True)
            set_cell(ws2, run_row, 3, r["tns"],          wf, center=True)
            set_cell(ws2, run_row, 4, r["total_power"],  BLU)
            set_cell(ws2, run_row, 5, r["cell_count"],   bg)
            set_cell(ws2, run_row, 6, r["utilization"],  bg, center=True)
            run_row += 1
        run_row += 1  # gap between corners

    for col, w in zip("ABCDEF", [16, 10, 10, 14, 12, 14]):
        ws2.column_dimensions[col].width = w

    # ────────────────────────────────────────────────────
    # Sheet 3: Compare Across Corners (group by run)
    # ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Compare Corners")
    runs = sorted(set(r["run"] for r in records))
    crow = 1
    for run in runs:
        run_records = [r for r in records if r["run"] == run]
        ws3.cell(crow, 1, f"Run: {run}").font = Font(bold=True, name="Arial", size=12, color="1F3864")
        crow += 1
        for ci, h in enumerate(["Corner", "WNS", "TNS", "Total Power", "Cell Count", "Utilization %"], 1):
            hdr(ws3, crow, ci, h)
        crow += 1
        for r in run_records:
            bg = rf(crow)
            wf = RED if (r["wns"] is not None and r["wns"] < 0) else GRN if r["wns"] is not None else bg
            set_cell(ws3, crow, 1, r["corner"],       bg)
            set_cell(ws3, crow, 2, r["wns"],          wf, bold=True, center=True)
            set_cell(ws3, crow, 3, r["tns"],          wf, center=True)
            set_cell(ws3, crow, 4, r["total_power"],  BLU)
            set_cell(ws3, crow, 5, r["cell_count"],   bg)
            set_cell(ws3, crow, 6, r["utilization"],  bg, center=True)
            crow += 1
        crow += 1

    for col, w in zip("ABCDEF", [20, 10, 10, 14, 12, 14]):
        ws3.column_dimensions[col].width = w

    # ────────────────────────────────────────────────────
    # Sheet 4: WNS Trend Chart (across runs)
    # ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("WNS Trend")
    ws4.cell(1, 1, "Run").font  = Font(bold=True, name="Arial")
    ws4.cell(1, 2, "WNS").font  = Font(bold=True, name="Arial")
    ws4.cell(1, 3, "Corner").font = Font(bold=True, name="Arial")
    for i, r in enumerate(records, 2):
        ws4.cell(i, 1, r["run"])
        ws4.cell(i, 2, r["wns"])
        ws4.cell(i, 3, r["corner"])

    if len(records) >= 2:
        chart = LineChart()
        chart.title  = "WNS Trend Across Runs"
        chart.y_axis.title = "WNS (ns)"
        chart.x_axis.title = "Run"
        chart.style  = 10
        chart.width  = 20
        chart.height = 12
        data = Reference(ws4, min_col=2, min_row=1, max_row=len(records)+1)
        cats = Reference(ws4, min_col=1, min_row=2, max_row=len(records)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws4.add_chart(chart, "E2")

    # ────────────────────────────────────────────────────
    # Sheet 5: Power Trend Chart
    # ────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Power Trend")
    for ci, h in enumerate(["Run", "Corner", "Total Power", "Dynamic", "Leakage"], 1):
        ws5.cell(1, ci, h).font = Font(bold=True, name="Arial")
    for i, r in enumerate(records, 2):
        ws5.cell(i, 1, r["run"])
        ws5.cell(i, 2, r["corner"])
        ws5.cell(i, 3, r["total_power"])
        ws5.cell(i, 4, r["dynamic"])
        ws5.cell(i, 5, r["leakage"])

    if len(records) >= 2:
        chart = BarChart()
        chart.title  = "Power Breakdown Across Runs"
        chart.y_axis.title = "Power (mW)"
        chart.x_axis.title = "Run"
        chart.style  = 10
        chart.width  = 20
        chart.height = 12
        chart.type   = "col"
        data = Reference(ws5, min_col=3, max_col=5, min_row=1, max_row=len(records)+1)
        cats = Reference(ws5, min_col=1, min_row=2, max_row=len(records)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws5.add_chart(chart, "G2")

    wb.save(EXCEL_FILE)
    print(f"\nExcel saved : {EXCEL_FILE}")
    print(f"  Sheet 1 → QoR Table       (all metrics, color coded)")
    print(f"  Sheet 2 → Compare Runs    (grouped by corner)")
    print(f"  Sheet 3 → Compare Corners (grouped by run)")
    print(f"  Sheet 4 → WNS Trend Chart")
    print(f"  Sheet 5 → Power Trend Chart")

# ── MAIN ────────────────────────────────────────────────
def main():
    files = pick_files()
    if not files:
        print("No files selected. Exiting.")
        return

    records = []
    for filepath in files:
        fname = os.path.basename(filepath)
        run, corner = get_run_info(fname)
        print(f"  Parsing: {fname} ...")
        data = parse_qor(filepath)
        data["run"]    = run
        data["corner"] = corner
        data["file"]   = fname
        records.append(data)

    print_terminal(records)
    save_excel(records)

if __name__ == "__main__":
    main()
