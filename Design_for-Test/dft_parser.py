import os
import re
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── CONFIG ──────────────────────────────────────────────
EXCEL_FILE = "dft_results.xlsx"

# ── STEP 1: Pick DFT report files ───────────────────────
def pick_files():
    print("\n=== VLSI DFT Report Parser ===")
    print("Parses: Fault coverage, scan chain, ATPG reports\n")

    folder = input("Enter folder path with DFT report files (or press Enter for current): ").strip()
    if not folder:
        folder = "."

    all_files = (
        glob.glob(os.path.join(folder, "*.rpt")) +
        glob.glob(os.path.join(folder, "*.log")) +
        glob.glob(os.path.join(folder, "*.txt"))
    )

    if not all_files:
        print(f"No report files found in: {folder}")
        return []

    print(f"\nFound {len(all_files)} file(s):\n")
    for i, f in enumerate(all_files):
        print(f"  [{i+1}] {os.path.basename(f)}")

    print("\n  A = All   or   1,2,3 = specific files")
    choice = input("\nYour choice: ").strip().upper()

    if choice == "A":
        return all_files
    selected = []
    for n in choice.split(","):
        n = n.strip()
        if n.isdigit() and 1 <= int(n) <= len(all_files):
            selected.append(all_files[int(n)-1])
    return selected

# ── STEP 2: Parse DFT report ────────────────────────────
def parse_dft(filepath):
    with open(filepath) as f:
        content = f.read()
        lines   = content.splitlines()

    data = {
        # Fault coverage
        "fault_coverage"    : None,
        "total_faults"      : None,
        "detected_faults"   : None,
        "undetected_faults" : None,
        "atpg_eff"          : None,
        # Scan chain
        "scan_chains"       : None,
        "scan_cells"        : None,
        "scan_length"       : [],
        # Patterns
        "total_patterns"    : None,
        "stuck_at_patterns" : None,
        "transition_patterns": None,
        # MBIST
        "mbist_coverage"    : None,
        "memories_tested"   : None,
    }

    scan_chains = []
    in_chain    = False

    for line in lines:
        s = line.strip()

        # ── Fault Coverage ──
        m = re.search(r"Fault\s+Coverage\s*[:\-]?\s*([\d.]+)\s*%?", s, re.IGNORECASE)
        if m: data["fault_coverage"] = float(m.group(1))

        m = re.search(r"Total\s+Faults?\s*[:\-]?\s*(\d+)", s, re.IGNORECASE)
        if m: data["total_faults"] = int(m.group(1))

        m = re.search(r"Detected\s+Faults?\s*[:\-]?\s*(\d+)", s, re.IGNORECASE)
        if m: data["detected_faults"] = int(m.group(1))

        m = re.search(r"(Undetected|UDID|Untestable)\s+Faults?\s*[:\-]?\s*(\d+)", s, re.IGNORECASE)
        if m: data["undetected_faults"] = int(m.group(2))

        m = re.search(r"ATPG\s+Eff\w*\s*[:\-]?\s*([\d.]+)\s*%?", s, re.IGNORECASE)
        if m: data["atpg_eff"] = float(m.group(1))

        # ── Scan Chains ──
        m = re.search(r"(Number of\s+)?Scan\s+Chains?\s*[:\-]?\s*(\d+)", s, re.IGNORECASE)
        if m: data["scan_chains"] = int(m.group(2))

        m = re.search(r"(Total\s+)?Scan\s+(Cells?|Flops?)\s*[:\-]?\s*(\d+)", s, re.IGNORECASE)
        if m: data["scan_cells"] = int(m.group(3))

        # Scan chain detail line:  "Chain 1 : length = 256"
        m = re.search(r"[Cc]hain\s+(\d+)\s*[:\-].*?[Ll]ength\s*[=:\-]?\s*(\d+)", s)
        if m:
            scan_chains.append({"chain": int(m.group(1)), "length": int(m.group(2))})

        # ── Patterns ──
        m = re.search(r"Total\s+[Pp]atterns?\s*[:\-]?\s*(\d+)", s)
        if m: data["total_patterns"] = int(m.group(1))

        m = re.search(r"Stuck.at\s+[Pp]atterns?\s*[:\-]?\s*(\d+)", s, re.IGNORECASE)
        if m: data["stuck_at_patterns"] = int(m.group(1))

        m = re.search(r"Transition\s+[Pp]atterns?\s*[:\-]?\s*(\d+)", s, re.IGNORECASE)
        if m: data["transition_patterns"] = int(m.group(1))

        # ── MBIST ──
        m = re.search(r"MBIST\s+Coverage\s*[:\-]?\s*([\d.]+)\s*%?", s, re.IGNORECASE)
        if m: data["mbist_coverage"] = float(m.group(1))

        m = re.search(r"Memories\s+[Tt]ested\s*[:\-]?\s*(\d+)", s, re.IGNORECASE)
        if m: data["memories_tested"] = int(m.group(1))

    data["scan_length"] = scan_chains

    # Compute undetected if missing
    if data["undetected_faults"] is None and data["total_faults"] and data["detected_faults"]:
        data["undetected_faults"] = data["total_faults"] - data["detected_faults"]

    return data

# ── STEP 3: Print terminal summary ──────────────────────
def print_terminal(fname, data):
    fc = f"{data['fault_coverage']}%" if data['fault_coverage'] else "-"
    status = "✅ PASS" if data['fault_coverage'] and data['fault_coverage'] >= 95 else "⚠️  CHECK"
    print(f"\n{'='*55}")
    print(f"  File            : {fname}")
    print(f"  Fault Coverage  : {fc}  {status}")
    print(f"  Total Faults    : {data['total_faults']}")
    print(f"  Detected        : {data['detected_faults']}")
    print(f"  Undetected      : {data['undetected_faults']}")
    print(f"  ATPG Efficiency : {data['atpg_eff']}%")
    print(f"  Scan Chains     : {data['scan_chains']}")
    print(f"  Scan Cells      : {data['scan_cells']}")
    print(f"  Total Patterns  : {data['total_patterns']}")
    print(f"  MBIST Coverage  : {data['mbist_coverage']}%")
    if data["scan_length"]:
        print(f"  Chain lengths   :")
        for c in data["scan_length"][:5]:
            print(f"    Chain {c['chain']} → {c['length']} cells")
    print(f"{'='*55}")

# ── STEP 4: Save to Excel ────────────────────────────────
def save_excel(all_results):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR = PatternFill("solid", start_color="1F3864")
    ALT = PatternFill("solid", start_color="F2F2F2")
    WHT = PatternFill("solid", start_color="FFFFFF")
    GRN = PatternFill("solid", start_color="C6EFCE")
    RED = PatternFill("solid", start_color="FFC7CE")
    YLW = PatternFill("solid", start_color="FFEB9C")
    BLU = PatternFill("solid", start_color="DDEEFF")

    def hdr(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill      = HDR
        c.alignment = Alignment(horizontal="center")

    def rf(r): return ALT if r % 2 == 0 else WHT

    def sc(ws, r, c, v, fill=None, bold=False, center=False):
        cell = ws.cell(r, c, v)
        cell.font = Font(bold=bold, name="Arial")
        if fill: cell.fill = fill
        if center: cell.alignment = Alignment(horizontal="center")

    def cov_fill(val):
        if val is None: return WHT
        if val >= 95:   return GRN
        if val >= 85:   return YLW
        return RED

    # ── Sheet 1: Fault Coverage Summary ──
    ws1 = wb.create_sheet("Fault Coverage")
    for ci, h in enumerate(["File","Fault Coverage %","Total Faults","Detected","Undetected","ATPG Eff %","Status"], 1):
        hdr(ws1, 1, ci, h)
    for col, w in zip("ABCDEFG", [28,16,13,12,13,12,10]):
        ws1.column_dimensions[col].width = w

    for ri, (fname, data) in enumerate(all_results, 2):
        f  = rf(ri)
        cf = cov_fill(data["fault_coverage"])
        status = "PASS" if data["fault_coverage"] and data["fault_coverage"] >= 95 else "CHECK"
        sf = GRN if status == "PASS" else YLW
        sc(ws1, ri, 1, fname,                   f)
        sc(ws1, ri, 2, data["fault_coverage"],  cf, bold=True, center=True)
        sc(ws1, ri, 3, data["total_faults"],    f,  center=True)
        sc(ws1, ri, 4, data["detected_faults"], GRN, center=True)
        sc(ws1, ri, 5, data["undetected_faults"], RED if data["undetected_faults"] else f, center=True)
        sc(ws1, ri, 6, data["atpg_eff"],        f,  center=True)
        sc(ws1, ri, 7, status,                  sf, bold=True, center=True)

    # ── Sheet 2: Scan Chain Summary ──
    ws2 = wb.create_sheet("Scan Chains")
    for ci, h in enumerate(["File","Scan Chains","Scan Cells","Total Patterns","Stuck-At","Transition"], 1):
        hdr(ws2, 1, ci, h)
    for col, w in zip("ABCDEF", [28,13,12,15,12,12]):
        ws2.column_dimensions[col].width = w

    for ri, (fname, data) in enumerate(all_results, 2):
        f = rf(ri)
        sc(ws2, ri, 1, fname,                      f)
        sc(ws2, ri, 2, data["scan_chains"],         BLU, center=True)
        sc(ws2, ri, 3, data["scan_cells"],          BLU, center=True)
        sc(ws2, ri, 4, data["total_patterns"],      f,   center=True)
        sc(ws2, ri, 5, data["stuck_at_patterns"],   f,   center=True)
        sc(ws2, ri, 6, data["transition_patterns"], f,   center=True)

    # ── Sheet 3: Scan Chain Detail (per chain length) ──
    ws3 = wb.create_sheet("Chain Detail")
    for ci, h in enumerate(["File","Chain #","Length (cells)"], 1):
        hdr(ws3, 1, ci, h)
    for col, w in zip("ABC", [28,10,15]):
        ws3.column_dimensions[col].width = w

    r = 2
    for fname, data in all_results:
        for ch in data["scan_length"]:
            f = rf(r)
            sc(ws3, r, 1, fname,          f)
            sc(ws3, r, 2, ch["chain"],    f, center=True)
            sc(ws3, r, 3, ch["length"],   BLU, bold=True, center=True)
            r += 1

    # ── Sheet 4: MBIST ──
    ws4 = wb.create_sheet("MBIST")
    for ci, h in enumerate(["File","MBIST Coverage %","Memories Tested"], 1):
        hdr(ws4, 1, ci, h)
    for col, w in zip("ABC", [28,18,16]):
        ws4.column_dimensions[col].width = w

    for ri, (fname, data) in enumerate(all_results, 2):
        f  = rf(ri)
        cf = cov_fill(data["mbist_coverage"])
        sc(ws4, ri, 1, fname,                   f)
        sc(ws4, ri, 2, data["mbist_coverage"],  cf, bold=True, center=True)
        sc(ws4, ri, 3, data["memories_tested"], f,  center=True)

    wb.save(EXCEL_FILE)
    print(f"\nExcel saved : {EXCEL_FILE}")
    print(f"  Sheet 1 → Fault Coverage  (coverage%, detected, undetected)")
    print(f"  Sheet 2 → Scan Chains     (chains, cells, patterns)")
    print(f"  Sheet 3 → Chain Detail    (per chain length)")
    print(f"  Sheet 4 → MBIST           (memory test coverage)")

# ── MAIN ────────────────────────────────────────────────
def main():
    files = pick_files()
    if not files:
        print("No files selected. Exiting.")
        return

    all_results = []
    for filepath in files:
        fname = os.path.basename(filepath)
        print(f"\nParsing: {fname} ...")
        data = parse_dft(filepath)
        print_terminal(fname, data)
        all_results.append((fname, data))

    save_excel(all_results)

if __name__ == "__main__":
    main()
