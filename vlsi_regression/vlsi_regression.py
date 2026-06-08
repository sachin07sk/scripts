import os
import glob
import subprocess
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill

# ── CONFIG ──────────────────────────────────────────────
EXCEL_FILE  = "regression_results.xlsx"   # Excel file to save results
SIMULATOR   = "python mock_vsim.py"       # Questa/ModelSim command
SIM_OPTIONS = "-c -do 'run -all; quit'"   # Run in batch mode and quit

# PASS / FAIL keywords to look for in log
PASS_KEYWORDS = ["TEST PASSED", "PASS", "Simulation complete"]
FAIL_KEYWORDS = ["TEST FAILED", "FAIL", "Error", "Fatal"]

# ── STEP 1: Ask user to pick test files ─────────────────
def pick_test_files():
    print("\n=== VLSI Regression Tool (Questa/ModelSim) ===\n")
    print("Supported file types: .v  .sv  .vhd\n")

    folder = input("Enter folder path where your test files are (or press Enter for current folder): ").strip()
    if not folder:
        folder = "."

    # Find all supported files
    all_files = (
        glob.glob(os.path.join(folder, "*.v"))  +
        glob.glob(os.path.join(folder, "*.sv")) +
        glob.glob(os.path.join(folder, "*.vhd"))
    )

    if not all_files:
        print(f"No .v / .sv / .vhd files found in: {folder}")
        return []

    print(f"\nFound {len(all_files)} test file(s):\n")
    for i, f in enumerate(all_files):
        print(f"  [{i+1}] {os.path.basename(f)}")

    print("\nOptions:")
    print("  A  = Run ALL files")
    print("  1,2,3 = Run specific files by number (comma separated)")
    choice = input("\nYour choice: ").strip().upper()

    if choice == "A":
        return all_files
    else:
        selected = []
        for num in choice.split(","):
            num = num.strip()
            if num.isdigit() and 1 <= int(num) <= len(all_files):
                selected.append(all_files[int(num) - 1])
        return selected

# ── STEP 2: Run one test and get PASS/FAIL ───────────────
def run_test(filepath):
    filename = os.path.basename(filepath)
    log_file = filename.replace(".", "_") + ".log"

    cmd = f"{SIMULATOR} {SIM_OPTIONS} {filepath} > {log_file} 2>&1"

    try:
        subprocess.run(cmd, shell=True, timeout=120)
        return parse_log(log_file), log_file
    except subprocess.TimeoutExpired:
        return "TIMEOUT", log_file
    except Exception as e:
        return "ERROR", log_file

def parse_log(log_file):
    if not os.path.exists(log_file):
        return "NO LOG"
    with open(log_file) as f:
        content = f.read()
    for kw in PASS_KEYWORDS:
        if kw in content:
            return "PASS"
    for kw in FAIL_KEYWORDS:
        if kw in content:
            return "FAIL"
    return "UNKNOWN"

# ── STEP 3: Save results to Excel ────────────────────────
def save_to_excel(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Regression Results"

    # Colors
    HEADER_FILL = PatternFill("solid", start_color="1F3864")
    PASS_FILL   = PatternFill("solid", start_color="C6EFCE")
    FAIL_FILL   = PatternFill("solid", start_color="FFC7CE")
    OTHER_FILL  = PatternFill("solid", start_color="FFEB9C")
    ALT_FILL    = PatternFill("solid", start_color="F2F2F2")

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30

    # Header row
    headers = ["#", "Test File Name", "Result", "Run Time", "Log File"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, (filepath, status, runtime, logfile) in enumerate(results, 2):
        filename = os.path.basename(filepath)
        fill = PASS_FILL if status == "PASS" else FAIL_FILL if status == "FAIL" else OTHER_FILL
        alt  = ALT_FILL if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")

        ws.cell(row=i, column=1, value=i-1).fill = alt
        ws.cell(row=i, column=2, value=filename).fill = alt
        cell_status = ws.cell(row=i, column=3, value=status)
        cell_status.fill      = fill
        cell_status.font      = Font(bold=True, name="Arial")
        cell_status.alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=4, value=runtime).fill = alt
        ws.cell(row=i, column=5, value=logfile).fill = alt

        for col in [1, 2, 4, 5]:
            ws.cell(row=i, column=col).font = Font(name="Arial")

    # Summary rows
    total   = len(results)
    passed  = sum(1 for r in results if r[1] == "PASS")
    failed  = total - passed
    gap_row = total + 2

    ws.cell(row=gap_row,   column=1, value="Total").font  = Font(bold=True, name="Arial")
    ws.cell(row=gap_row,   column=2, value=f'=COUNTA(B2:B{total+1})').font = Font(name="Arial")
    ws.cell(row=gap_row+1, column=1, value="Passed").font = Font(bold=True, name="Arial", color="375623")
    ws.cell(row=gap_row+1, column=2, value=f'=COUNTIF(C2:C{total+1},"PASS")').font = Font(name="Arial")
    ws.cell(row=gap_row+2, column=1, value="Failed").font = Font(bold=True, name="Arial", color="9C0006")
    ws.cell(row=gap_row+2, column=2, value=f'=COUNTIF(C2:C{total+1},"FAIL")').font = Font(name="Arial")
    ws.cell(row=gap_row+3, column=1, value="Pass Rate").font = Font(bold=True, name="Arial")
    ws.cell(row=gap_row+3, column=2, value=f'=B{gap_row+1}/B{gap_row}').number_format = "0.0%"
    ws.cell(row=gap_row+3, column=2).font = Font(name="Arial")

    wb.save(EXCEL_FILE)
    print(f"\nResults saved to: {EXCEL_FILE}")

# ── MAIN ─────────────────────────────────────────────────
def main():
    test_files = pick_test_files()

    if not test_files:
        print("No files selected. Exiting.")
        return

    print(f"\nRunning {len(test_files)} test(s)...\n")

    results = []
    for filepath in test_files:
        name = os.path.basename(filepath)
        print(f"  Testing: {name} ...", end="", flush=True)
        start   = datetime.datetime.now()
        status, logfile = run_test(filepath)
        elapsed = str(datetime.datetime.now() - start).split(".")[0]
        print(f"  {status}")
        results.append((filepath, status, elapsed, logfile))

    save_to_excel(results)

    # Quick terminal summary
    passed = sum(1 for r in results if r[1] == "PASS")
    print(f"\n  PASSED : {passed}/{len(results)}")
    print(f"  FAILED : {len(results)-passed}/{len(results)}\n")

if __name__ == "__main__":
    main()
