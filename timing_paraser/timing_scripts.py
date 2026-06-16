import os
import re
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── CONFIG ──────────────────────────────────────────────
EXCEL_FILE = "timing_results.xlsx"

# ── STEP 1: Ask user to pick report files ───────────────
def pick_report_files():
    print("\n=== VLSI Timing Report Parser ===\n")
    print("Supported: Questa / PrimeTime / Innovus timing reports (.rpt .txt .log)\n")

    folder = input("Enter folder path with timing reports (or press Enter for current): ").strip()
    if not folder:
        folder = "."

    all_files = (
        glob.glob(os.path.join(folder, "*.rpt")) +
        glob.glob(os.path.join(folder, "*.txt")) +
        glob.glob(os.path.join(folder, "*.log"))
    )

    if not all_files:
        print(f"No report files found in: {folder}")
        return []

    print(f"\nFound {len(all_files)} file(s):\n")
    for i, f in enumerate(all_files):
        print(f"  [{i+1}] {os.path.basename(f)}")

    print("\n  A = Parse ALL   or   1,2,3 = specific files")
    choice = input("\nYour choice: ").strip().upper()

    if choice == "A":
        return all_files
    else:
        selected = []
        for num in choice.split(","):
            num = num.strip()
            if num.isdigit() and 1 <= int(num) <= len(all_files):
                selected.append(all_files[int(num) - 1])
        return selected

# ── STEP 2: Parse one timing report ─────────────────────
def parse_report(filepath):
    with open(filepath) as f:
        content = f.read()
        lines   = content.splitlines()

    paths = []
    current = None

    for i, line in enumerate(lines):

        # ── Detect start of a timing path block ──
        # Matches:  "Path 1", "Startpoint:", or "slack (VIOLATED)"
        if re.search(r"^Path\s+\d+", line) or "Startpoint:" in line:
            if current:
                paths.append(current)
            current = {
                "startpoint" : "",
                "endpoint"   : "",
                "path_group" : "",
                "slack"      : None,
                "status"     : "",
                "cells"      : []
            }

        if current is None:
            continue

        # ── Startpoint ──
        m = re.search(r"Startpoint\s*[:\-]?\s*(.+)", line)
        if m:
            current["startpoint"] = m.group(1).strip()

        # ── Endpoint ──
        m = re.search(r"Endpoint\s*[:\-]?\s*(.+)", line)
        if m:
            current["endpoint"] = m.group(1).strip()

        # ── Path group ──
        m = re.search(r"Path Group\s*[:\-]?\s*(.+)", line)
        if m:
            current["path_group"] = m.group(1).strip()

        # ── Slack value ──
        # Matches:  "slack (VIOLATED) -0.35"  or  "slack (MET) 0.12"
        m = re.search(r"slack\s*\((VIOLATED|MET)\)\s*([-\d.]+)", line)
        if m:
            current["status"] = m.group(1)
            current["slack"]  = float(m.group(2))

        # Also catch plain: "slack = -0.35"
        if current["slack"] is None:
            m = re.search(r"slack\s*=\s*([-\d.]+)", line)
            if m:
                current["slack"] = float(m.group(1))
                current["status"] = "VIOLATED" if float(m.group(1)) < 0 else "MET"

        # ── Cell / instance lines ──
        # Matches lines with cell name + arrival time pattern
        # Example:  "  U_ADD/A    AND2     0.23   0.45"
        m = re.search(r"^\s{2,}(\S+)\s+(\S+)\s+([-\d.]+)\s+([-\d.]+)", line)
        if m:
            current["cells"].append({
                "instance" : m.group(1),
                "cell_type": m.group(2),
                "incr"     : m.group(3),
                "path_delay": m.group(4)
            })

    # Append last path
    if current:
        paths.append(current)

    # ── Compute WNS and TNS ──
    slacks = [p["slack"] for p in paths if p["slack"] is not None]
    wns = min(slacks) if slacks else None
    tns = sum(s for s in slacks if s < 0) if slacks else None

    return paths, wns, tns

# ── STEP 3: Print summary to terminal ───────────────────
def print_terminal(filename, paths, wns, tns):
    violated = [p for p in paths if p["status"] == "VIOLATED"]
    print(f"\n{'='*55}")
    print(f"  File    : {filename}")
    print(f"  Paths   : {len(paths)}   Violated: {len(violated)}")
    print(f"  WNS     : {wns}")
    print(f"  TNS     : {round(tns, 4) if tns else 0}")
    print(f"{'='*55}")

    for i, p in enumerate(paths, 1):
        status_str = "❌ VIOLATED" if p["status"] == "VIOLATED" else "✅ MET"
        print(f"\n  Path {i} | {status_str} | Slack: {p['slack']}")
        print(f"    From : {p['startpoint']}")
        print(f"    To   : {p['endpoint']}")
        print(f"    Group: {p['path_group']}")
        if p["cells"]:
            print(f"    Cells ({len(p['cells'])}):")
            for c in p["cells"][:5]:   # show first 5 cells
                print(f"      {c['instance']:<30} {c['cell_type']:<15} delay={c['path_delay']}")
            if len(p["cells"]) > 5:
                print(f"      ... and {len(p['cells'])-5} more cells")

# ── STEP 4: Save to Excel ────────────────────────────────
def save_excel(all_results):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──
    ws_sum = wb.active
    ws_sum.title = "Summary"

    HDR_FILL  = PatternFill("solid", start_color="1F3864")
    PASS_FILL = PatternFill("solid", start_color="C6EFCE")
    FAIL_FILL = PatternFill("solid", start_color="FFC7CE")
    ALT_FILL  = PatternFill("solid", start_color="F2F2F2")

    def hdr(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill      = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    sum_headers = ["File", "Total Paths", "Violated", "MET", "WNS", "TNS"]
    for ci, h in enumerate(sum_headers, 1):
        hdr(ws_sum, 1, ci, h)

    ws_sum.column_dimensions["A"].width = 30
    for col in ["B","C","D","E","F"]:
        ws_sum.column_dimensions[col].width = 14

    for ri, (fname, paths, wns, tns) in enumerate(all_results, 2):
        violated = sum(1 for p in paths if p["status"] == "VIOLATED")
        met      = len(paths) - violated
        fill     = FAIL_FILL if violated > 0 else PASS_FILL
        alt      = ALT_FILL if ri % 2 == 0 else PatternFill("solid", start_color="FFFFFF")

        ws_sum.cell(ri, 1, fname).fill = alt
        ws_sum.cell(ri, 2, len(paths)).fill = alt
        c = ws_sum.cell(ri, 3, violated)
        c.fill = fill; c.font = Font(bold=True, name="Arial")
        c.alignment = Alignment(horizontal="center")
        ws_sum.cell(ri, 4, met).fill = alt
        ws_sum.cell(ri, 5, wns).fill = alt
        ws_sum.cell(ri, 6, round(tns, 4) if tns else 0).fill = alt
        for col in [1,2,4,5,6]:
            ws_sum.cell(ri, col).font = Font(name="Arial")

    # ── Sheet 2: All Paths detail ──
    ws_paths = wb.create_sheet("Path Details")
    path_headers = ["File", "Path#", "Status", "Slack", "Startpoint", "Endpoint", "Path Group", "Cell Count"]
    for ci, h in enumerate(path_headers, 1):
        hdr(ws_paths, 1, ci, h)

    for col, w in zip(["A","B","C","D","E","F","G","H"], [25,7,10,10,35,35,20,10]):
        ws_paths.column_dimensions[col].width = w

    row = 2
    for fname, paths, wns, tns in all_results:
        for pi, p in enumerate(paths, 1):
            fill = FAIL_FILL if p["status"] == "VIOLATED" else PASS_FILL
            alt  = ALT_FILL if row % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
            ws_paths.cell(row, 1, fname).fill = alt
            ws_paths.cell(row, 2, pi).fill = alt
            c = ws_paths.cell(row, 3, p["status"])
            c.fill = fill; c.font = Font(bold=True, name="Arial")
            c.alignment = Alignment(horizontal="center")
            ws_paths.cell(row, 4, p["slack"]).fill = alt
            ws_paths.cell(row, 5, p["startpoint"]).fill = alt
            ws_paths.cell(row, 6, p["endpoint"]).fill = alt
            ws_paths.cell(row, 7, p["path_group"]).fill = alt
            ws_paths.cell(row, 8, len(p["cells"])).fill = alt
            for col in [1,2,4,5,6,7,8]:
                ws_paths.cell(row, col).font = Font(name="Arial")
            row += 1

    # ── Sheet 3: Cell details ──
    ws_cells = wb.create_sheet("Cell Details")
    cell_headers = ["File", "Path#", "Slack", "Instance", "Cell Type", "Incr Delay", "Path Delay"]
    for ci, h in enumerate(cell_headers, 1):
        hdr(ws_cells, 1, ci, h)
    for col, w in zip(["A","B","C","D","E","F","G"], [25,7,10,35,15,12,12]):
        ws_cells.column_dimensions[col].width = w

    row = 2
    for fname, paths, wns, tns in all_results:
        for pi, p in enumerate(paths, 1):
            for c in p["cells"]:
                alt = ALT_FILL if row % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
                ws_cells.cell(row, 1, fname).fill = alt
                ws_cells.cell(row, 2, pi).fill = alt
                ws_cells.cell(row, 3, p["slack"]).fill = alt
                ws_cells.cell(row, 4, c["instance"]).fill = alt
                ws_cells.cell(row, 5, c["cell_type"]).fill = alt
                ws_cells.cell(row, 6, c["incr"]).fill = alt
                ws_cells.cell(row, 7, c["path_delay"]).fill = alt
                for col in range(1, 8):
                    ws_cells.cell(row, col).font = Font(name="Arial")
                row += 1

    wb.save(EXCEL_FILE)
    print(f"\nExcel saved: {EXCEL_FILE}")
    print(f"  Sheet 1 → Summary (WNS / TNS per file)")
    print(f"  Sheet 2 → Path Details (all paths)")
    print(f"  Sheet 3 → Cell Details (every cell in every path)")

# ── MAIN ─────────────────────────────────────────────────
def main():
    files = pick_report_files()
    if not files:
        print("No files selected. Exiting.")
        return

    all_results = []
    for filepath in files:
        fname = os.path.basename(filepath)
        print(f"\nParsing: {fname} ...")
        paths, wns, tns = parse_report(filepath)
        print_terminal(fname, paths, wns, tns)
        all_results.append((fname, paths, wns, tns))

    save_excel(all_results)

if __name__ == "__main__":
    main()
