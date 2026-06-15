import os
import subprocess
import datetime
import time
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_FILE  = "job_results.xlsx"
LOG_DIR     = "./job_logs"
MAX_PARALLEL = 4    # Max jobs running at same time

def pick_jobs():
    print("\n=== VLSI Job Scheduler ===")
    print("Submit multiple jobs in parallel (LSF/Grid/local)\n")

    folder = input("Folder with job scripts (.tcl .sh .py) (or press Enter for current): ").strip()
    if not folder: folder = "."

    all_files = (glob.glob(os.path.join(folder,"*.tcl")) +
                 glob.glob(os.path.join(folder,"*.sh"))  +
                 glob.glob(os.path.join(folder,"*.py")))

    if not all_files:
        print(f"No job scripts found in {folder}"); return []

    print(f"\nFound {len(all_files)} job script(s):\n")
    for i,f in enumerate(all_files): print(f"  [{i+1}] {os.path.basename(f)}")
    print("\n  A = All   or   1,2,3 = specific")
    choice = input("\nYour choice: ").strip().upper()

    mode = input("\nRun mode — L=Local  B=LSF bsub  G=Grid qsub  (default: L): ").strip().upper() or "L"

    if choice == "A": return all_files, mode
    sel = []
    for n in choice.split(","):
        n=n.strip()
        if n.isdigit() and 1<=int(n)<=len(all_files): sel.append(all_files[int(n)-1])
    return sel, mode

def build_cmd(script, mode):
    ext = os.path.splitext(script)[1]
    if mode == "B":      # LSF bsub
        return f"bsub -q normal -n 4 -R 'rusage[mem=8000]' '{script}'"
    elif mode == "G":    # Grid qsub
        return f"qsub -pe smp 4 -l h_vmem=8G '{script}'"
    else:                # Local
        if ext == ".sh":  return f"bash '{script}'"
        elif ext == ".py": return f"python3 '{script}'"
        elif ext == ".tcl": return f"tclsh '{script}'"
        return f"'{script}'"

def run_jobs(scripts, mode):
    os.makedirs(LOG_DIR, exist_ok=True)
    results  = []
    running  = []
    pending  = list(scripts)

    print(f"\nSubmitting {len(scripts)} jobs  |  max parallel: {MAX_PARALLEL}\n")

    while pending or running:
        # Launch new jobs up to MAX_PARALLEL
        while pending and len(running) < MAX_PARALLEL:
            script = pending.pop(0)
            fname  = os.path.basename(script)
            log    = os.path.join(LOG_DIR, fname + ".log")
            cmd    = build_cmd(script, mode)
            start  = datetime.datetime.now()
            try:
                proc = subprocess.Popen(cmd, shell=True,
                                        stdout=open(log,"w"), stderr=subprocess.STDOUT)
                running.append({"script": fname, "proc": proc, "start": start, "log": log})
                print(f"  [SUBMITTED] {fname}  PID={proc.pid}")
            except Exception as e:
                results.append({"script": fname, "status": "ERROR", "time": "-", "log": log})
                print(f"  [ERROR]     {fname}  → {e}")

        # Check running jobs
        still_running = []
        for job in running:
            ret = job["proc"].poll()
            if ret is None:
                still_running.append(job)
            else:
                elapsed = str(datetime.datetime.now() - job["start"]).split(".")[0]
                status  = "PASS" if ret == 0 else "FAIL"
                color   = "\033[92m" if status=="PASS" else "\033[91m"
                print(f"  [DONE]      {job['script']}  {color}{status}\033[0m  ({elapsed})")
                results.append({"script": job["script"], "status": status, "time": elapsed, "log": job["log"]})
        running = still_running

        if running: time.sleep(2)

    return results

def save_excel(results):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Job Results"
    HDR=PatternFill("solid",start_color="1F3864"); GRN=PatternFill("solid",start_color="C6EFCE")
    RED=PatternFill("solid",start_color="FFC7CE"); ALT=PatternFill("solid",start_color="F2F2F2")
    WHT=PatternFill("solid",start_color="FFFFFF")

    for ci,h in enumerate(["#","Job Script","Status","Runtime","Log File"],1):
        c=ws.cell(1,ci,h); c.font=Font(bold=True,color="FFFFFF",name="Arial")
        c.fill=HDR; c.alignment=Alignment(horizontal="center")
    for col,w in zip("ABCDE",[5,35,10,12,40]): ws.column_dimensions[col].width=w

    passed=failed=0
    for ri,r in enumerate(results,2):
        bg=ALT if ri%2==0 else WHT
        sf=GRN if r["status"]=="PASS" else RED
        if r["status"]=="PASS": passed+=1
        else: failed+=1
        ws.cell(ri,1,ri-1).fill=bg; ws.cell(ri,1).font=Font(name="Arial")
        ws.cell(ri,2,r["script"]).fill=bg; ws.cell(ri,2).font=Font(name="Arial")
        c=ws.cell(ri,3,r["status"]); c.fill=sf; c.font=Font(bold=True,name="Arial"); c.alignment=Alignment(horizontal="center")
        ws.cell(ri,4,r["time"]).fill=bg; ws.cell(ri,4).font=Font(name="Arial")
        ws.cell(ri,5,r["log"]).fill=bg; ws.cell(ri,5).font=Font(name="Arial")

    gap=len(results)+3
    for label,val in [("Total",len(results)),("Passed",passed),("Failed",failed)]:
        ws.cell(gap,1,label).font=Font(bold=True,name="Arial")
        ws.cell(gap,2,val).font=Font(name="Arial")
        gap+=1

    wb.save(EXCEL_FILE)
    print(f"\n  Job report saved: {EXCEL_FILE}")
    print(f"  Passed: {passed}  Failed: {failed}  Total: {len(results)}")

def main():
    result = pick_jobs()
    if not result: return
    scripts, mode = result
    if not scripts: print("No scripts selected."); return
    results = run_jobs(scripts, mode)
    save_excel(results)

if __name__ == "__main__":
    main()
