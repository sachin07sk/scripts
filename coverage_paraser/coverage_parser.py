import os
import re
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_FILE = "coverage_results.xlsx"

def pick_files():
    print("\n=== VLSI Coverage Report Parser ===")
    print("Parses: Code, Functional, Toggle, FSM coverage\n")
    folder = input("Enter folder path with coverage reports (or press Enter for current): ").strip()
    if not folder: folder = "."
    all_files = (glob.glob(os.path.join(folder,"*.rpt")) +
                 glob.glob(os.path.join(folder,"*.log")) +
                 glob.glob(os.path.join(folder,"*.txt")))
    if not all_files:
        print(f"No files found in: {folder}"); return []
    print(f"\nFound {len(all_files)} file(s):\n")
    for i,f in enumerate(all_files): print(f"  [{i+1}] {os.path.basename(f)}")
    print("\n  A = All   or   1,2,3 = specific")
    choice = input("\nYour choice: ").strip().upper()
    if choice == "A": return all_files
    selected = []
    for n in choice.split(","):
        n = n.strip()
        if n.isdigit() and 1 <= int(n) <= len(all_files):
            selected.append(all_files[int(n)-1])
    return selected

def parse_coverage(filepath):
    with open(filepath) as f: content = f.read()
    data = {
        "line_coverage"       : None,
        "branch_coverage"     : None,
        "condition_coverage"  : None,
        "toggle_coverage"     : None,
        "fsm_coverage"        : None,
        "functional_coverage" : None,
        "total_bins"          : None,
        "covered_bins"        : None,
        "uncovered_bins"      : None,
        "modules"             : []
    }
    for line in content.splitlines():
        s = line.strip()
        m = re.search(r"Line\s+[Cc]overage\s*[:\-]?\s*([\d.]+)\s*%?", s)
        if m: data["line_coverage"] = float(m.group(1))
        m = re.search(r"[Bb]ranch\s+[Cc]overage\s*[:\-]?\s*([\d.]+)\s*%?", s)
        if m: data["branch_coverage"] = float(m.group(1))
        m = re.search(r"[Cc]ondition\s+[Cc]overage\s*[:\-]?\s*([\d.]+)\s*%?", s)
        if m: data["condition_coverage"] = float(m.group(1))
        m = re.search(r"[Tt]oggle\s+[Cc]overage\s*[:\-]?\s*([\d.]+)\s*%?", s)
        if m: data["toggle_coverage"] = float(m.group(1))
        m = re.search(r"FSM\s+[Cc]overage\s*[:\-]?\s*([\d.]+)\s*%?", s)
        if m: data["fsm_coverage"] = float(m.group(1))
        m = re.search(r"[Ff]unctional\s+[Cc]overage\s*[:\-]?\s*([\d.]+)\s*%?", s)
        if m: data["functional_coverage"] = float(m.group(1))
        m = re.search(r"[Tt]otal\s+[Bb]ins?\s*[:\-]?\s*(\d+)", s)
        if m: data["total_bins"] = int(m.group(1))
        m = re.search(r"[Cc]overed\s+[Bb]ins?\s*[:\-]?\s*(\d+)", s)
        if m: data["covered_bins"] = int(m.group(1))
        m = re.search(r"[Uu]ncovered\s+[Bb]ins?\s*[:\-]?\s*(\d+)", s)
        if m: data["uncovered_bins"] = int(m.group(1))
        # Module level: "module_name   95.5%"
        m = re.search(r"^(\w+)\s+([\d.]+)\s*%", s)
        if m and float(m.group(2)) <= 100:
            data["modules"].append({"name": m.group(1), "coverage": float(m.group(2))})
    if data["uncovered_bins"] is None and data["total_bins"] and data["covered_bins"]:
        data["uncovered_bins"] = data["total_bins"] - data["covered_bins"]
    return data

def print_terminal(fname, data):
    print(f"\n{'='*55}")
    print(f"  File               : {fname}")
    print(f"  Line Coverage      : {data['line_coverage']}%")
    print(f"  Branch Coverage    : {data['branch_coverage']}%")
    print(f"  Condition Coverage : {data['condition_coverage']}%")
    print(f"  Toggle Coverage    : {data['toggle_coverage']}%")
    print(f"  FSM Coverage       : {data['fsm_coverage']}%")
    print(f"  Functional Coverage: {data['functional_coverage']}%")
    print(f"  Total Bins         : {data['total_bins']}   Covered: {data['covered_bins']}   Uncovered: {data['uncovered_bins']}")
    print(f"{'='*55}")

def save_excel(all_results):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    HDR = PatternFill("solid", start_color="1F3864")
    ALT = PatternFill("solid", start_color="F2F2F2")
    WHT = PatternFill("solid", start_color="FFFFFF")
    GRN = PatternFill("solid", start_color="C6EFCE")
    RED = PatternFill("solid", start_color="FFC7CE")
    YLW = PatternFill("solid", start_color="FFEB9C")

    def hdr(ws,row,col,val):
        c=ws.cell(row=row,column=col,value=val)
        c.font=Font(bold=True,color="FFFFFF",name="Arial",size=11)
        c.fill=HDR; c.alignment=Alignment(horizontal="center")
    def rf(r): return ALT if r%2==0 else WHT
    def sc(ws,r,c,v,fill=None,bold=False,center=False):
        cell=ws.cell(r,c,v); cell.font=Font(bold=bold,name="Arial")
        if fill: cell.fill=fill
        if center: cell.alignment=Alignment(horizontal="center")
    def cf(v):
        if v is None: return WHT
        if v>=95: return GRN
        if v>=85: return YLW
        return RED

    ws1=wb.create_sheet("Coverage Summary")
    for ci,h in enumerate(["File","Line %","Branch %","Condition %","Toggle %","FSM %","Functional %"],1): hdr(ws1,1,ci,h)
    for col,w in zip("ABCDEFG",[28,10,10,13,10,9,14]): ws1.column_dimensions[col].width=w
    for ri,(fname,data) in enumerate(all_results,2):
        f=rf(ri)
        sc(ws1,ri,1,fname,f)
        for ci,key in enumerate(["line_coverage","branch_coverage","condition_coverage","toggle_coverage","fsm_coverage","functional_coverage"],2):
            sc(ws1,ri,ci,data[key],cf(data[key]),bold=True,center=True)

    ws2=wb.create_sheet("Bins")
    for ci,h in enumerate(["File","Total Bins","Covered","Uncovered"],1): hdr(ws2,1,ci,h)
    for col,w in zip("ABCD",[28,12,12,12]): ws2.column_dimensions[col].width=w
    for ri,(fname,data) in enumerate(all_results,2):
        f=rf(ri); unv=data["uncovered_bins"] or 0
        sc(ws2,ri,1,fname,f)
        sc(ws2,ri,2,data["total_bins"],f,center=True)
        sc(ws2,ri,3,data["covered_bins"],GRN,center=True)
        sc(ws2,ri,4,unv,RED if unv else GRN,bold=bool(unv),center=True)

    ws3=wb.create_sheet("Module Coverage")
    for ci,h in enumerate(["File","Module Name","Coverage %"],1): hdr(ws3,1,ci,h)
    for col,w in zip("ABC",[28,30,12]): ws3.column_dimensions[col].width=w
    r=2
    for fname,data in all_results:
        for mod in data["modules"]:
            f=rf(r)
            sc(ws3,r,1,fname,f); sc(ws3,r,2,mod["name"],f)
            sc(ws3,r,3,mod["coverage"],cf(mod["coverage"]),bold=True,center=True)
            r+=1

    wb.save(EXCEL_FILE)
    print(f"\nExcel saved : {EXCEL_FILE}")
    print(f"  Sheet 1 → Coverage Summary (line, branch, toggle, FSM)")
    print(f"  Sheet 2 → Bins             (total, covered, uncovered)")
    print(f"  Sheet 3 → Module Coverage  (per module %)")

def main():
    files = pick_files()
    if not files: print("No files selected. Exiting."); return
    all_results = []
    for filepath in files:
        fname = os.path.basename(filepath)
        print(f"\nParsing: {fname} ...")
        data = parse_coverage(filepath)
        print_terminal(fname, data)
        all_results.append((fname, data))
    save_excel(all_results)

if __name__ == "__main__":
    main()
