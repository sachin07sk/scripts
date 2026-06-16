import os
import re
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── CONFIG ──────────────────────────────────────────────
EXCEL_FILE = "power_results.xlsx"

# ── STEP 1: Pick log files ───────────────────────────────
def pick_files():
    print("\n=== VLSI Power Analysis Parser (Questa Power) ===\n")

    folder = input("Enter folder path with power log files (or press Enter for current): ").strip()
    if not folder:
        folder = "."

    all_files = glob.glob(os.path.join(folder, "*.log"))

    if not all_files:
        print(f"No .log files found in: {folder}")
        return []

    print(f"\nFound {len(all_files)} file(s):\n")
    for i, f in enumerate(all_files):
        print(f"  [{i+1}] {os.path.basename(f)}")

    print("\n  A = All   or   1,2,3 = specific files")
    choice = input("\nYour choice: ").strip().upper()

    if choice == "A":
        return all_files
    selected = []
    for n in choice.split(","):
        n = n.strip()
        if n.isdigit() and 1 <= int(n) <= len(all_files):
            selected.append(all_files[int(n)-1])
    return selected

# ── STEP 2: Parse Questa Power log ──────────────────────
def parse_questa_power(filepath):
    with open(filepath) as f:
        lines = f.readlines()

    summary = {
        "dynamic"  : None,   # Switching power
        "internal" : None,   # Internal (glitch) power
        "leakage"  : None,   # Static leakage power
        "total"    : None,   # Total power
        "unit"     : "mW"
    }
    cells = []
    in_cell_section = False

    for line in lines:
        s = line.strip()

        # ── Detect power unit ──
        m = re.search(r"Power Unit\s*[:\-]?\s*(\S+)", s, re.IGNORECASE)
        if m:
            summary["unit"] = m.group(1)

        # ── Summary power values ──
        # Questa formats: "Dynamic Power   :   12.34 mW"
        #             or  "Switching Power    12.34"
        m = re.search(r"(Switching|Dynamic)\s+Power\s*[:\-]?\s*([\d.eE+\-]+)", s, re.IGNORECASE)
        if m:
            summary["dynamic"] = float(m.group(2))

        m = re.search(r"Internal\s+Power\s*[:\-]?\s*([\d.eE+\-]+)", s, re.IGNORECASE)
        if m:
            summary["internal"] = float(m.group(1))

        m = re.search(r"(Leakage|Static)\s+Power\s*[:\-]?\s*([\d.eE+\-]+)", s, re.IGNORECASE)
        if m:
            summary["leakage"] = float(m.group(2))

        m = re.search(r"Total\s+Power\s*[:\-]?\s*([\d.eE+\-]+)", s, re.IGNORECASE)
        if m:
            summary["total"] = float(m.group(1))

        # ── Per-cell section detection ──
        # Questa prints a table header like:
        # "Instance   Cell    Dynamic  Internal  Leakage  Total"
        if re.search(r"Instance\s+.*(Dynamic|Power)", s, re.IGNORECASE):
            in_cell_section = True
            continue

        if in_cell_section:
            # Skip separator lines (--- or ===)
            if re.match(r"^[-=]+$", s) or s == "":
                continue
            # Stop at end of table
            if re.match(r"^(END|Total|Summary)", s, re.IGNORECASE):
                in_cell_section = False
                continue

            # Parse cell row
            # Format: instance_name  cell_type  dyn  internal  leak  total
            parts = s.split()
            if len(parts) >= 4:
                try:
                    # Try to extract up to 6 columns
                    cell = {
                        "instance" : parts[0],
                        "cell_type": parts[1] if len(parts) >= 6 else "",
                        "dynamic"  : float(parts[-4]) if len(parts) >= 5 else None,
                        "internal" : float(parts[-3]) if len(parts) >= 4 else None,
                        "leakage"  : float(parts[-2]) if len(parts) >= 4 else None,
                        "total"    : float(parts[-1])
                    }
                    cells.append(cell)
                except ValueError:
                    pass  # skip malformed lines

    # If total not found, calculate from components
    if summary["total"] is None:
        parts = [summary["dynamic"], summary["internal"], summary["leakage"]]
        parts = [p for p in parts if p is not None]
        if parts:
            summary["total"] = round(sum(parts), 6)

    return summary, cells

# ── STEP 3: Print terminal summary ──────────────────────
def print_terminal(fname, summary, cells):
    unit = summary["unit"]
    print(f"\n{'='*55}")
    print(f"  File     : {fname}")
    print(f"  Dynamic  : {summary['dynamic']}  {unit}")
    print(f"  Internal : {summary['internal']}  {unit}")
    print(f"  Leakage  : {summary['leakage']}  {unit}")
    print(f"  Total    : {summary['total']}  {unit}")
    print(f"  Cells    : {len(cells)}")
    if cells:
        print(f"\n  Top 5 power consumers:")
        top5 = sorted([c for c in cells if c["total"] is not None], key=lambda x: x["total"], reverse=True)[:5]
        for c in top5:
            print(f"    {c['instance']:<35} {c['cell_type']:<15} total={c['total']} {unit}")
    print(f"{'='*55}")

# ── STEP 4: Save to Excel ────────────────────────────────
def save_excel(all_results):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR  = PatternFill("solid", start_color="1F3864")
    ALT  = PatternFill("solid", start_color="F2F2F2")
    WHT  = PatternFill("solid", start_color="FFFFFF")
    HIGH = PatternFill("solid", start_color="FFC7CE")   # high power = red
    MED  = PatternFill("solid", start_color="FFEB9C")   # medium     = yellow
    LOW  = PatternFill("solid", start_color="C6EFCE")   # low        = green

    def hdr(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill      = HDR
        c.alignment = Alignment(horizontal="center")

    def rf(r):
        return ALT if r % 2 == 0 else WHT

    def power_fill(val, max_val):
        if val is None or max_val == 0:
            return WHT
        ratio = val / max_val
        if ratio > 0.6:   return HIGH
        if ratio > 0.3:   return MED
        return LOW

    # ── Sheet 1: Power Summary ──
    ws = wb.create_sheet("Power Summary")
    for ci, h in enumerate(["File", "Dynamic (mW)", "Internal (mW)", "Leakage (mW)", "Total (mW)", "Unit"], 1):
        hdr(ws, 1, ci, h)
    for col, w in zip("ABCDEF", [30, 15, 15, 15, 15, 10]):
        ws.column_dimensions[col].width = w

    r = 2
    for fname, summary, cells in all_results:
        f = rf(r)
        for ci, v in enumerate([fname, summary["dynamic"], summary["internal"],
                                  summary["leakage"], summary["total"], summary["unit"]], 1):
            ws.cell(r, ci, v).fill = f
            ws.cell(r, ci).font = Font(name="Arial")
        # Highlight total cell
        tot_cell = ws.cell(r, 5)
        tot_cell.font = Font(bold=True, name="Arial")
        r += 1

    # ── Sheet 2: Per-Cell Breakdown ──
    ws = wb.create_sheet("Cell Breakdown")
    for ci, h in enumerate(["File", "Instance", "Cell Type", "Dynamic", "Internal", "Leakage", "Total"], 1):
        hdr(ws, 1, ci, h)
    for col, w in zip("ABCDEFG", [25, 35, 20, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    r = 2
    for fname, summary, cells in all_results:
        if not cells:
            continue
        max_total = max((c["total"] for c in cells if c["total"] is not None), default=0)
        # Sort by total power descending
        sorted_cells = sorted(cells, key=lambda x: x["total"] if x["total"] else 0, reverse=True)
        for c in sorted_cells:
            pf = power_fill(c["total"], max_total)
            bg = rf(r)
            ws.cell(r, 1, fname).fill = bg
            ws.cell(r, 1).font = Font(name="Arial")
            ws.cell(r, 2, c["instance"]).fill = bg
            ws.cell(r, 2).font = Font(name="Arial")
            ws.cell(r, 3, c["cell_type"]).fill = bg
            ws.cell(r, 3).font = Font(name="Arial")
            for ci, key in enumerate(["dynamic", "internal", "leakage", "total"], 4):
                cell = ws.cell(r, ci, c[key])
                cell.fill = pf if key == "total" else bg
                cell.font = Font(bold=(key=="total"), name="Arial")
            r += 1

    # ── Sheet 3: Top Power Consumers (across all files) ──
    ws = wb.create_sheet("Top Consumers")
    for ci, h in enumerate(["File", "Instance", "Cell Type", "Total Power", "% of Chip"], 1):
        hdr(ws, 1, ci, h)
    for col, w in zip("ABCDE", [25, 35, 20, 15, 12]):
        ws.column_dimensions[col].width = w

    all_cells_flat = []
    for fname, summary, cells in all_results:
        chip_total = summary["total"] or 1
        for c in cells:
            if c["total"] is not None:
                all_cells_flat.append((fname, c, chip_total))

    top20 = sorted(all_cells_flat, key=lambda x: x[1]["total"], reverse=True)[:20]
    for r_idx, (fname, c, chip_total) in enumerate(top20, 2):
        pct  = round((c["total"] / chip_total) * 100, 2) if chip_total else 0
        pf   = HIGH if pct > 10 else MED if pct > 5 else LOW
        bg   = rf(r_idx)
        ws.cell(r_idx, 1, fname).fill = bg
        ws.cell(r_idx, 1).font = Font(name="Arial")
        ws.cell(r_idx, 2, c["instance"]).fill = bg
        ws.cell(r_idx, 2).font = Font(name="Arial")
        ws.cell(r_idx, 3, c["cell_type"]).fill = bg
        ws.cell(r_idx, 3).font = Font(name="Arial")
        cell = ws.cell(r_idx, 4, c["total"])
        cell.fill = pf; cell.font = Font(bold=True, name="Arial")
        ws.cell(r_idx, 5, f"{pct}%").fill = pf
        ws.cell(r_idx, 5).font = Font(name="Arial")

    wb.save(EXCEL_FILE)
    print(f"\nExcel saved : {EXCEL_FILE}")
    print(f"  Sheet 1 → Power Summary  (dynamic, internal, leakage, total)")
    print(f"  Sheet 2 → Cell Breakdown (all cells sorted by power)")
    print(f"  Sheet 3 → Top Consumers  (top 20 cells across all files)")

# ── MAIN ────────────────────────────────────────────────
def main():
    files = pick_files()
    if not files:
        print("No files selected. Exiting.")
        return

    all_results = []
    for filepath in files:
        fname = os.path.basename(filepath)
        print(f"\nParsing: {fname} ...")
        summary, cells = parse_questa_power(filepath)
        print_terminal(fname, summary, cells)
        all_results.append((fname, summary, cells))

    save_excel(all_results)

if __name__ == "__main__":
    main()
