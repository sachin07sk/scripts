import os
import re
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── CONFIG ──────────────────────────────────────────────
EXCEL_FILE = "netlist_results.xlsx"

# ── STEP 1: Pick netlist files ──────────────────────────
def pick_files():
    print("\n=== VLSI Netlist Parser ===")
    print("Supported: Gate-level netlists (.v .sv .vhd)\n")

    folder = input("Enter folder path with netlist files (or press Enter for current): ").strip()
    if not folder:
        folder = "."

    all_files = (
        glob.glob(os.path.join(folder, "*.v"))  +
        glob.glob(os.path.join(folder, "*.sv")) +
        glob.glob(os.path.join(folder, "*.vhd"))
    )

    if not all_files:
        print(f"No netlist files found in: {folder}")
        return []

    print(f"\nFound {len(all_files)} file(s):\n")
    for i, f in enumerate(all_files):
        print(f"  [{i+1}] {os.path.basename(f)}")

    print("\n  A = All   or   1,2,3 = specific files")
    choice = input("\nYour choice: ").strip().upper()

    if choice == "A":
        return all_files
    selected = []
    for n in choice.split(","):
        n = n.strip()
        if n.isdigit() and 1 <= int(n) <= len(all_files):
            selected.append(all_files[int(n)-1])
    return selected

# ── STEP 2: Parse netlist ────────────────────────────────
def parse_netlist(filepath):
    with open(filepath) as f:
        content = f.read()

    # Remove comments
    content = re.sub(r"//.*", "", content)
    content = re.sub(r"/\*.*?\*/", "", content, flags=re.DOTALL)

    modules   = []
    ports     = []
    wires     = []
    instances = []
    cells     = {}   # cell_type -> count

    # ── Modules ──
    for m in re.finditer(r"module\s+(\w+)\s*[#(]", content):
        modules.append(m.group(1))

    # ── Ports ──
    for m in re.finditer(r"\b(input|output|inout)\s+(?:wire\s+)?(?:\[[\d\s:\-]+\]\s+)?(\w+)", content):
        ports.append({"direction": m.group(1), "name": m.group(2)})

    # ── Wires / regs ──
    for m in re.finditer(r"\b(wire|reg)\s+(?:\[[\d\s:\-]+\]\s+)?(\w+)", content):
        wires.append({"type": m.group(1), "name": m.group(2)})

    # ── Cell instances ──
    # Pattern: CELL_TYPE  instance_name ( .port(wire), ... );
    for m in re.finditer(r"^\s*(\w+)\s+(\w+)\s*\(", content, re.MULTILINE):
        cell_type = m.group(1)
        inst_name = m.group(2)
        # Skip keywords
        if cell_type in ("module","input","output","inout","wire","reg","assign","always","begin","end","if","else","case"):
            continue
        instances.append({"cell_type": cell_type, "inst_name": inst_name})
        cells[cell_type] = cells.get(cell_type, 0) + 1

    # ── Assigns ──
    assign_count = len(re.findall(r"\bassign\b", content))

    return {
        "modules"      : modules,
        "ports"        : ports,
        "wires"        : wires,
        "instances"    : instances,
        "cells"        : cells,
        "assign_count" : assign_count,
        "total_cells"  : len(instances),
        "unique_cells" : len(cells),
        "input_ports"  : sum(1 for p in ports if p["direction"] == "input"),
        "output_ports" : sum(1 for p in ports if p["direction"] == "output"),
        "inout_ports"  : sum(1 for p in ports if p["direction"] == "inout"),
    }

# ── STEP 3: Print terminal summary ──────────────────────
def print_terminal(fname, data):
    print(f"\n{'='*55}")
    print(f"  File         : {fname}")
    print(f"  Modules      : {len(data['modules'])} → {', '.join(data['modules'][:5])}")
    print(f"  Total cells  : {data['total_cells']}")
    print(f"  Unique cells : {data['unique_cells']}")
    print(f"  Ports        : IN={data['input_ports']}  OUT={data['output_ports']}  INOUT={data['inout_ports']}")
    print(f"  Wires/Regs   : {len(data['wires'])}")
    print(f"  Assigns      : {data['assign_count']}")
    print(f"\n  Top 5 cell types:")
    top5 = sorted(data["cells"].items(), key=lambda x: x[1], reverse=True)[:5]
    for cell, cnt in top5:
        print(f"    {cell:<25} count = {cnt}")
    print(f"{'='*55}")

# ── STEP 4: Save to Excel ────────────────────────────────
def save_excel(all_results):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR = PatternFill("solid", start_color="1F3864")
    ALT = PatternFill("solid", start_color="F2F2F2")
    WHT = PatternFill("solid", start_color="FFFFFF")
    BLU = PatternFill("solid", start_color="DDEEFF")
    GRN = PatternFill("solid", start_color="C6EFCE")

    def hdr(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill      = HDR
        c.alignment = Alignment(horizontal="center")

    def rf(r): return ALT if r % 2 == 0 else WHT

    def sc(ws, r, c, v, fill=None, bold=False, center=False):
        cell = ws.cell(r, c, v)
        cell.font = Font(bold=bold, name="Arial")
        if fill: cell.fill = fill
        if center: cell.alignment = Alignment(horizontal="center")

    # ── Sheet 1: Summary ──
    ws1 = wb.create_sheet("Summary")
    for ci, h in enumerate(["File","Modules","Total Cells","Unique Cells","Inputs","Outputs","Inouts","Wires","Assigns"], 1):
        hdr(ws1, 1, ci, h)
    for col, w in zip("ABCDEFGHI", [28,10,12,13,8,9,8,8,9]):
        ws1.column_dimensions[col].width = w

    for ri, (fname, data) in enumerate(all_results, 2):
        f = rf(ri)
        sc(ws1, ri, 1, fname,                  f)
        sc(ws1, ri, 2, len(data["modules"]),   f, center=True)
        sc(ws1, ri, 3, data["total_cells"],    BLU, bold=True, center=True)
        sc(ws1, ri, 4, data["unique_cells"],   f, center=True)
        sc(ws1, ri, 5, data["input_ports"],    GRN, center=True)
        sc(ws1, ri, 6, data["output_ports"],   GRN, center=True)
        sc(ws1, ri, 7, data["inout_ports"],    f, center=True)
        sc(ws1, ri, 8, len(data["wires"]),     f, center=True)
        sc(ws1, ri, 9, data["assign_count"],   f, center=True)

    # ── Sheet 2: Cell Count (all types) ──
    ws2 = wb.create_sheet("Cell Count")
    for ci, h in enumerate(["File","Cell Type","Count","% of Total"], 1):
        hdr(ws2, 1, ci, h)
    for col, w in zip("ABCD", [28,25,10,12]):
        ws2.column_dimensions[col].width = w

    r = 2
    for fname, data in all_results:
        total = data["total_cells"] or 1
        sorted_cells = sorted(data["cells"].items(), key=lambda x: x[1], reverse=True)
        for cell_type, count in sorted_cells:
            f = rf(r)
            pct = round((count / total) * 100, 1)
            sc(ws2, r, 1, fname,      f)
            sc(ws2, r, 2, cell_type,  f)
            sc(ws2, r, 3, count,      BLU, bold=True, center=True)
            sc(ws2, r, 4, f"{pct}%",  f, center=True)
            r += 1

    # ── Sheet 3: Ports ──
    ws3 = wb.create_sheet("Ports")
    for ci, h in enumerate(["File","Port Name","Direction"], 1):
        hdr(ws3, 1, ci, h)
    for col, w in zip("ABC", [28,30,12]):
        ws3.column_dimensions[col].width = w

    r = 2
    dir_fill = {"input": GRN, "output": BLU, "inout": ALT}
    for fname, data in all_results:
        for p in data["ports"]:
            f = rf(r)
            sc(ws3, r, 1, fname,         f)
            sc(ws3, r, 2, p["name"],     f)
            sc(ws3, r, 3, p["direction"], dir_fill.get(p["direction"], f), center=True)
            r += 1

    # ── Sheet 4: All Instances ──
    ws4 = wb.create_sheet("Instances")
    for ci, h in enumerate(["File","Instance Name","Cell Type"], 1):
        hdr(ws4, 1, ci, h)
    for col, w in zip("ABC", [28,35,25]):
        ws4.column_dimensions[col].width = w

    r = 2
    for fname, data in all_results:
        for inst in data["instances"]:
            f = rf(r)
            sc(ws4, r, 1, fname,              f)
            sc(ws4, r, 2, inst["inst_name"],  f)
            sc(ws4, r, 3, inst["cell_type"],  f)
            r += 1

    wb.save(EXCEL_FILE)
    print(f"\nExcel saved : {EXCEL_FILE}")
    print(f"  Sheet 1 → Summary      (cell count, ports, wires)")
    print(f"  Sheet 2 → Cell Count   (every cell type + count + %)")
    print(f"  Sheet 3 → Ports        (all IO ports + direction)")
    print(f"  Sheet 4 → Instances    (every instance + cell type)")

# ── MAIN ────────────────────────────────────────────────
def main():
    files = pick_files()
    if not files:
        print("No files selected. Exiting.")
        return

    all_results = []
    for filepath in files:
        fname = os.path.basename(filepath)
        print(f"\nParsing: {fname} ...")
        data = parse_netlist(filepath)
        print_terminal(fname, data)
        all_results.append((fname, data))

    save_excel(all_results)

if __name__ == "__main__":
    main()
